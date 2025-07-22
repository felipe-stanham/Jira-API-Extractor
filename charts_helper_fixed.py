"""Fixed helper functions for creating charts in Excel with proper formula syntax."""

from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from collections import Counter

def create_fixed_charts_sheet(wb, issues, worklogs=None):
    """Creates a comprehensive charts sheet with properly formatted Excel formulas."""
    ws_charts = wb.create_sheet(title="Charts")
    
    if not issues:
        ws_charts.append(['No data available for charts'])
        return
    
    # ===== ISSUES BY STATUS CHART =====
    ws_charts.append(['Issues by Status Analysis'])
    ws_charts.append(['Status', 'Count'])
    
    # Get unique statuses from Sprint Issues sheet
    if 'Sprint Issues' in wb.sheetnames:
        ws_issues = wb['Sprint Issues']
        
        # Get unique statuses
        unique_statuses = set()
        for row in range(2, ws_issues.max_row + 1):
            status = ws_issues.cell(row=row, column=4).value  # Status is in column D
            if status:
                unique_statuses.add(status)
        
        status_start_row = 3
        for i, status in enumerate(sorted(unique_statuses)):
            row = status_start_row + i
            ws_charts.cell(row=row, column=1, value=status)
            # Use proper sheet name escaping for formulas
            formula = f'=COUNTIF(\'Sprint Issues\'.D:D,A{row})'
            ws_charts.cell(row=row, column=2, value=formula)
        
        status_end_row = status_start_row + len(unique_statuses) - 1
        
        # Create pie chart for Issues by Status
        pie_status = PieChart()
        pie_status.title = "Issues by Status"
        pie_status.width = 12
        pie_status.height = 8
        
        labels = Reference(ws_charts, min_col=1, min_row=status_start_row, max_row=status_end_row)
        data = Reference(ws_charts, min_col=2, min_row=status_start_row, max_row=status_end_row)
        
        pie_status.add_data(data, titles_from_data=False)
        pie_status.set_categories(labels)
        
        # Configure chart appearance
        pie_status.dataLabels = DataLabelList()
        pie_status.dataLabels.showCatName = True
        pie_status.dataLabels.showVal = True
        pie_status.dataLabels.showPercent = True
        
        ws_charts.add_chart(pie_status, "D2")
    
    # ===== ISSUES BY TYPE CHART =====
    type_start_row = status_end_row + 3 if 'Sprint Issues' in wb.sheetnames else 3
    ws_charts.cell(row=type_start_row - 1, column=1, value='Issues by Type Analysis')
    ws_charts.cell(row=type_start_row, column=1, value='Issue Type')
    ws_charts.cell(row=type_start_row, column=2, value='Count')
    
    if 'Sprint Issues' in wb.sheetnames:
        ws_issues = wb['Sprint Issues']
        
        # Get unique issue types
        unique_types = set()
        for row in range(2, ws_issues.max_row + 1):
            issue_type = ws_issues.cell(row=row, column=2).value  # Issue Type is in column B
            if issue_type:
                unique_types.add(issue_type)
        
        type_data_start = type_start_row + 1
        for i, issue_type in enumerate(sorted(unique_types)):
            row = type_data_start + i
            ws_charts.cell(row=row, column=1, value=issue_type)
            # Use cell reference instead of hardcoded string
            formula = f'=COUNTIF(\'Sprint Issues\'.B:B,A{row})'
            ws_charts.cell(row=row, column=2, value=formula)
        
        type_end_row = type_data_start + len(unique_types) - 1
        
        # Create pie chart for Issues by Type
        pie_type = PieChart()
        pie_type.title = "Issues by Type"
        pie_type.width = 12
        pie_type.height = 8
        
        labels_type = Reference(ws_charts, min_col=1, min_row=type_data_start, max_row=type_end_row)
        data_type = Reference(ws_charts, min_col=2, min_row=type_data_start, max_row=type_end_row)
        
        pie_type.add_data(data_type, titles_from_data=False)
        pie_type.set_categories(labels_type)
        
        pie_type.dataLabels = DataLabelList()
        pie_type.dataLabels.showCatName = True
        pie_type.dataLabels.showVal = True
        pie_type.dataLabels.showPercent = True
        
        ws_charts.add_chart(pie_type, "D15")
    
    # ===== TIME BY ISSUE TYPE CHART =====
    if worklogs and 'Work Logs' in wb.sheetnames:
        time_start_row = type_end_row + 3 if 'Sprint Issues' in wb.sheetnames else 3
        ws_charts.cell(row=time_start_row - 1, column=1, value='Total Time by Issue Type')
        ws_charts.cell(row=time_start_row, column=1, value='Issue Type')
        ws_charts.cell(row=time_start_row, column=2, value='Total Hours')
        
        ws_worklogs = wb['Work Logs']
        
        # Get unique issue types from worklogs
        unique_worklog_types = set()
        for row in range(2, ws_worklogs.max_row + 1):
            issue_type = ws_worklogs.cell(row=row, column=2).value  # Issue Type is in column B
            if issue_type:
                unique_worklog_types.add(issue_type)
        
        time_data_start = time_start_row + 1
        for i, issue_type in enumerate(sorted(unique_worklog_types)):
            row = time_data_start + i
            ws_charts.cell(row=row, column=1, value=issue_type)
            # Use cell reference for SUMIF
            formula = f'=SUMIF(\'Work Logs\'.B:B,A{row},\'Work Logs\'.G:G)'
            ws_charts.cell(row=row, column=2, value=formula)
        
        time_end_row = time_data_start + len(unique_worklog_types) - 1
        
        # Create pie chart for Time by Issue Type
        pie_time = PieChart()
        pie_time.title = "Total Time by Issue Type"
        pie_time.width = 12
        pie_time.height = 8
        
        labels_time = Reference(ws_charts, min_col=1, min_row=time_data_start, max_row=time_end_row)
        data_time = Reference(ws_charts, min_col=2, min_row=time_data_start, max_row=time_end_row)
        
        pie_time.add_data(data_time, titles_from_data=False)
        pie_time.set_categories(labels_time)
        
        pie_time.dataLabels = DataLabelList()
        pie_time.dataLabels.showCatName = True
        pie_time.dataLabels.showVal = True
        pie_time.dataLabels.showPercent = True
        
        ws_charts.add_chart(pie_time, "D28")
        
        # ===== TIME BY AUTHOR AND ISSUE TYPE BAR CHART =====
        author_start_row = time_end_row + 3
        ws_charts.cell(row=author_start_row - 1, column=1, value='Time Spent by Author and Issue Type')
        
        # Get unique authors and issue types
        unique_authors = set()
        unique_worklog_types = set()
        
        for row in range(2, ws_worklogs.max_row + 1):
            author = ws_worklogs.cell(row=row, column=5).value  # Author is in column E
            issue_type = ws_worklogs.cell(row=row, column=2).value  # Issue Type is in column B
            if author:
                unique_authors.add(author)
            if issue_type:
                unique_worklog_types.add(issue_type)
        
        unique_authors = sorted(unique_authors)
        unique_worklog_types = sorted(unique_worklog_types)
        
        # Create header row
        header_row = author_start_row
        ws_charts.cell(row=header_row, column=1, value='Author')
        
        col = 2
        for issue_type in unique_worklog_types:
            ws_charts.cell(row=header_row, column=col, value=issue_type)
            col += 1
        ws_charts.cell(row=header_row, column=col, value='Total')
        
        # Create data rows
        data_start_row = author_start_row + 1
        for i, author in enumerate(unique_authors):
            row = data_start_row + i
            ws_charts.cell(row=row, column=1, value=author)
            
            col = 2
            for issue_type in unique_worklog_types:
                # Use cell references for SUMIFS
                author_cell = f'$A{row}'
                type_cell = f'{ws_charts.cell(row=header_row, column=col).coordinate}'
                formula = f'=SUMIFS(\'Work Logs\'.G:G,\'Work Logs\'.E:E,{author_cell},\'Work Logs\'.B:B,{type_cell})'
                ws_charts.cell(row=row, column=col, value=formula)
                col += 1
            
            # Total for author using cell reference
            total_formula = f'=SUMIF(\'Work Logs\'.E:E,{author_cell},\'Work Logs\'.G:G)'
            ws_charts.cell(row=row, column=col, value=total_formula)
        
        author_end_row = data_start_row + len(unique_authors) - 1
        
        # Create simplified bar chart without series titles to avoid compatibility issues
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "Time Spent by Author and Issue Type"
        bar_chart.y_axis.title = 'Hours'
        bar_chart.x_axis.title = 'Authors'
        bar_chart.width = 20
        bar_chart.height = 12
        
        # Add data series for each issue type
        for i, issue_type in enumerate(unique_worklog_types):
            data_col = 2 + i
            data_ref = Reference(ws_charts, min_col=data_col, min_row=data_start_row, max_row=author_end_row)
            bar_chart.add_data(data_ref, titles_from_data=False)
        
        # Add total series
        total_col = 2 + len(unique_worklog_types)
        total_ref = Reference(ws_charts, min_col=total_col, min_row=data_start_row, max_row=author_end_row)
        bar_chart.add_data(total_ref, titles_from_data=False)
        
        # Set categories (authors)
        cats = Reference(ws_charts, min_col=1, min_row=data_start_row, max_row=author_end_row)
        bar_chart.set_categories(cats)
        
        # Don't set series titles to avoid compatibility issues
        
        ws_charts.add_chart(bar_chart, "D41")
        
        # Add summary information with formulas
        summary_row = author_end_row + 3
        ws_charts.cell(row=summary_row, column=1, value='Summary Information')
        ws_charts.cell(row=summary_row + 1, column=1, value='Total Issues:')
        if 'Sprint Issues' in wb.sheetnames:
            ws_charts.cell(row=summary_row + 1, column=2, value='=COUNTA(\'Sprint Issues\'.A:A)-1')
        
        ws_charts.cell(row=summary_row + 2, column=1, value='Total Work Log Entries:')
        ws_charts.cell(row=summary_row + 2, column=2, value='=COUNTA(\'Work Logs\'.A:A)-1')
        ws_charts.cell(row=summary_row + 3, column=1, value='Total Hours Logged:')
        ws_charts.cell(row=summary_row + 3, column=2, value='=SUM(\'Work Logs\'.G:G)')
        
        if 'Comments' in wb.sheetnames:
            ws_charts.cell(row=summary_row + 4, column=1, value='Total Comments:')
            ws_charts.cell(row=summary_row + 4, column=2, value='=COUNTA(\'Comments\'.A:A)-1')
    else:
        # Add summary for issues only
        summary_row = type_end_row + 3
        ws_charts.cell(row=summary_row, column=1, value='Summary Information')
        ws_charts.cell(row=summary_row + 1, column=1, value='Total Issues:')
        if 'Sprint Issues' in wb.sheetnames:
            ws_charts.cell(row=summary_row + 1, column=2, value='=COUNTA(\'Sprint Issues\'.A:A)-1')
