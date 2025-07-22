"""Clean helper functions for creating charts in Excel without complex formulas."""

from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from collections import Counter

def create_clean_charts_sheet(wb, issues, worklogs=None):
    """Creates a charts sheet with static data to ensure Excel compatibility."""
    ws_charts = wb.create_sheet(title="Charts")
    
    if not issues:
        ws_charts.append(['No data available for charts'])
        return
    
    # ===== ISSUES BY STATUS CHART =====
    ws_charts.append(['Issues by Status Analysis'])
    ws_charts.append(['Status', 'Count'])
    
    # Count issues by status using Python (no Excel formulas)
    status_counts = Counter()
    for issue in issues:
        status = issue.get('fields', {}).get('status', {}).get('name', 'Unknown')
        status_counts[status] += 1
    
    status_start_row = 3
    for i, (status, count) in enumerate(sorted(status_counts.items())):
        row = status_start_row + i
        ws_charts.cell(row=row, column=1, value=status)
        ws_charts.cell(row=row, column=2, value=count)
    
    status_end_row = status_start_row + len(status_counts) - 1
    
    # Create pie chart for Issues by Status
    pie_status = PieChart()
    pie_status.title = "Issues by Status"
    pie_status.width = 12
    pie_status.height = 8
    
    labels = Reference(ws_charts, min_col=1, min_row=status_start_row, max_row=status_end_row)
    data = Reference(ws_charts, min_col=2, min_row=status_start_row, max_row=status_end_row)
    
    pie_status.add_data(data, titles_from_data=False)
    pie_status.set_categories(labels)
    
    # Configure chart appearance
    pie_status.dataLabels = DataLabelList()
    pie_status.dataLabels.showCatName = True
    pie_status.dataLabels.showVal = True
    pie_status.dataLabels.showPercent = True
    
    ws_charts.add_chart(pie_status, "D2")
    
    # ===== ISSUES BY TYPE CHART =====
    type_start_row = status_end_row + 3
    ws_charts.cell(row=type_start_row - 1, column=1, value='Issues by Type Analysis')
    ws_charts.cell(row=type_start_row, column=1, value='Issue Type')
    ws_charts.cell(row=type_start_row, column=2, value='Count')
    
    # Count issues by type using Python
    type_counts = Counter()
    for issue in issues:
        issue_type = issue.get('fields', {}).get('issuetype', {}).get('name', 'Unknown')
        type_counts[issue_type] += 1
    
    type_data_start = type_start_row + 1
    for i, (issue_type, count) in enumerate(sorted(type_counts.items())):
        row = type_data_start + i
        ws_charts.cell(row=row, column=1, value=issue_type)
        ws_charts.cell(row=row, column=2, value=count)
    
    type_end_row = type_data_start + len(type_counts) - 1
    
    # Create pie chart for Issues by Type
    pie_type = PieChart()
    pie_type.title = "Issues by Type"
    pie_type.width = 12
    pie_type.height = 8
    
    labels_type = Reference(ws_charts, min_col=1, min_row=type_data_start, max_row=type_end_row)
    data_type = Reference(ws_charts, min_col=2, min_row=type_data_start, max_row=type_end_row)
    
    pie_type.add_data(data_type, titles_from_data=False)
    pie_type.set_categories(labels_type)
    
    pie_type.dataLabels = DataLabelList()
    pie_type.dataLabels.showCatName = True
    pie_type.dataLabels.showVal = True
    pie_type.dataLabels.showPercent = True
    
    ws_charts.add_chart(pie_type, "D15")
    
    # ===== TIME BY ISSUE TYPE CHART =====
    if worklogs:
        time_start_row = type_end_row + 3
        ws_charts.cell(row=time_start_row - 1, column=1, value='Total Time by Issue Type')
        ws_charts.cell(row=time_start_row, column=1, value='Issue Type')
        ws_charts.cell(row=time_start_row, column=2, value='Total Hours')
        
        # Calculate time by issue type using Python
        time_by_type = {}
        for log in worklogs:
            issue_type = log.get('issueType', 'Unknown')
            hours = log.get('timeSpentHours', 0)
            if issue_type in time_by_type:
                time_by_type[issue_type] += hours
            else:
                time_by_type[issue_type] = hours
        
        time_data_start = time_start_row + 1
        for i, (issue_type, total_hours) in enumerate(sorted(time_by_type.items())):
            row = time_data_start + i
            ws_charts.cell(row=row, column=1, value=issue_type)
            ws_charts.cell(row=row, column=2, value=round(total_hours, 2))
        
        time_end_row = time_data_start + len(time_by_type) - 1
        
        # Create pie chart for Time by Issue Type
        pie_time = PieChart()
        pie_time.title = "Total Time by Issue Type"
        pie_time.width = 12
        pie_time.height = 8
        
        labels_time = Reference(ws_charts, min_col=1, min_row=time_data_start, max_row=time_end_row)
        data_time = Reference(ws_charts, min_col=2, min_row=time_data_start, max_row=time_end_row)
        
        pie_time.add_data(data_time, titles_from_data=False)
        pie_time.set_categories(labels_time)
        
        pie_time.dataLabels = DataLabelList()
        pie_time.dataLabels.showCatName = True
        pie_time.dataLabels.showVal = True
        pie_time.dataLabels.showPercent = True
        
        ws_charts.add_chart(pie_time, "D28")
        
        # ===== SIMPLE BAR CHART FOR TIME BY AUTHOR =====
        author_start_row = time_end_row + 3
        ws_charts.cell(row=author_start_row - 1, column=1, value='Total Time by Author')
        ws_charts.cell(row=author_start_row, column=1, value='Author')
        ws_charts.cell(row=author_start_row, column=2, value='Total Hours')
        
        # Calculate time by author using Python
        time_by_author = {}
        for log in worklogs:
            author = log.get('author', 'Unknown')
            hours = log.get('timeSpentHours', 0)
            if author in time_by_author:
                time_by_author[author] += hours
            else:
                time_by_author[author] = hours
        
        author_data_start = author_start_row + 1
        for i, (author, total_hours) in enumerate(sorted(time_by_author.items())):
            row = author_data_start + i
            ws_charts.cell(row=row, column=1, value=author)
            ws_charts.cell(row=row, column=2, value=round(total_hours, 2))
        
        author_end_row = author_data_start + len(time_by_author) - 1
        
        # Create simple bar chart
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "Total Time by Author"
        bar_chart.y_axis.title = 'Hours'
        bar_chart.x_axis.title = 'Authors'
        bar_chart.width = 15
        bar_chart.height = 10
        
        # Add single data series
        data_ref = Reference(ws_charts, min_col=2, min_row=author_data_start, max_row=author_end_row)
        bar_chart.add_data(data_ref, titles_from_data=False)
        
        # Set categories (authors)
        cats = Reference(ws_charts, min_col=1, min_row=author_data_start, max_row=author_end_row)
        bar_chart.set_categories(cats)
        
        ws_charts.add_chart(bar_chart, "D41")
        
        # Add summary information (static values, no formulas)
        summary_row = author_end_row + 3
        ws_charts.cell(row=summary_row, column=1, value='Summary Information')
        ws_charts.cell(row=summary_row + 1, column=1, value='Total Issues:')
        ws_charts.cell(row=summary_row + 1, column=2, value=len(issues))
        
        ws_charts.cell(row=summary_row + 2, column=1, value='Total Work Log Entries:')
        ws_charts.cell(row=summary_row + 2, column=2, value=len(worklogs))
        
        total_hours = sum(log.get('timeSpentHours', 0) for log in worklogs)
        ws_charts.cell(row=summary_row + 3, column=1, value='Total Hours Logged:')
        ws_charts.cell(row=summary_row + 3, column=2, value=round(total_hours, 2))
    else:
        # Add summary for issues only
        summary_row = type_end_row + 3
        ws_charts.cell(row=summary_row, column=1, value='Summary Information')
        ws_charts.cell(row=summary_row + 1, column=1, value='Total Issues:')
        ws_charts.cell(row=summary_row + 1, column=2, value=len(issues))
