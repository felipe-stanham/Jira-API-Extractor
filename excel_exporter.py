"""Excel export functionality for Jira data."""

from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table, TableStyleInfo
from collections import Counter
from charts_helper_enhanced import create_clean_charts_sheet
from config import JIRA_STORY_POINTS_FIELD
from progress_data_aggregator import calculate_epic_progress, calculate_sprint_composition
from progress_charts_helper import create_percentage_bar_chart, create_stacked_bar_chart, create_composition_pie_chart

class ExcelExporter:
    """Handles Excel export functionality."""
    
    def __init__(self):
        self.wb = None
    
    def save_to_excel(self, issues, worklogs, comments, filename="JiraExport.xlsx", issues_by_sprint=None, epic_label_issues=None, open_epic_issues=None):
        """Saves the fetched data to an Excel file with separate sheets and charts."""
        self.wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])
        
        sheets_created = []
        
        # Create Sprint Issues sheets - one for each sprint if multiple sprints provided
        if issues_by_sprint:
            for sprint_id, sprint_data in issues_by_sprint.items():
                sprint_issues = sprint_data['issues']
                sprint_name = sprint_data['name']
                
                # Create sheet with sprint name (truncate if too long for Excel)
                sheet_title = f"Sprint {sprint_id}"
                if len(sheet_title) > 31:  # Excel sheet name limit
                    sheet_title = f"Sprint {sprint_id}"[:31]
                
                ws_issues = self.wb.create_sheet(title=sheet_title)
                ws_issues.append(['Issue Key', 'Issue Type', 'Summary', 'Status', 'Sprint', 'Parent Summary', 'Story Points', 'Parent Key', 'Status Category'])
                
                for issue in sprint_issues:
                    parent_summary = 'N/A'
                    parent_key = 'N/A'
                    parent_field = issue.get('fields', {}).get('parent')
                    if parent_field:
                        parent_summary = parent_field.get('fields', {}).get('summary', 'N/A')
                        parent_key = parent_field.get('key', 'N/A')
                    
                    # Get story points (use 0 if not set)
                    story_points = issue.get('fields', {}).get(JIRA_STORY_POINTS_FIELD)
                    if story_points is None:
                        story_points = 0
                    
                    # Get status category
                    status_category = issue.get('fields', {}).get('status', {}).get('statusCategory', {}).get('name', 'N/A')
                    
                    ws_issues.append([
                        issue.get('key'),
                        issue.get('fields', {}).get('issuetype', {}).get('name', 'N/A'),
                        issue.get('fields', {}).get('summary'),
                        issue.get('fields', {}).get('status', {}).get('name', 'N/A'),
                        sprint_name,
                        parent_summary,
                        story_points,
                        parent_key,
                        status_category
                    ])
                
                sheets_created.append(sheet_title)
        elif issues:
            # Fallback for single sprint or backward compatibility
            ws_issues = self.wb.create_sheet(title="Sprint Issues")
            # Check if issues have sprint information
            has_sprint_info = any(issue.get('sprint_name') for issue in issues)
            
            if has_sprint_info:
                ws_issues.append(['Issue Key', 'Issue Type', 'Summary', 'Status', 'Sprint', 'Parent Summary', 'Story Points', 'Parent Key', 'Status Category'])
            else:
                ws_issues.append(['Issue Key', 'Issue Type', 'Summary', 'Status', 'Parent Summary', 'Story Points', 'Parent Key', 'Status Category'])
            
            for issue in issues:
                parent_summary = 'N/A'
                parent_key = 'N/A'
                parent_field = issue.get('fields', {}).get('parent')
                if parent_field:
                    parent_summary = parent_field.get('fields', {}).get('summary', 'N/A')
                    parent_key = parent_field.get('key', 'N/A')
                
                # Get story points (use 0 if not set)
                story_points = issue.get('fields', {}).get(JIRA_STORY_POINTS_FIELD)
                if story_points is None:
                    story_points = 0
                
                # Get status category
                status_category = issue.get('fields', {}).get('status', {}).get('statusCategory', {}).get('name', 'N/A')
                
                row_data = [
                    issue.get('key'),
                    issue.get('fields', {}).get('issuetype', {}).get('name', 'N/A'),
                    issue.get('fields', {}).get('summary'),
                    issue.get('fields', {}).get('status', {}).get('name', 'N/A')
                ]
                
                if has_sprint_info:
                    row_data.append(issue.get('sprint_name', ''))
                
                row_data.extend([parent_summary, story_points, parent_key, status_category])
                ws_issues.append(row_data)
            
            sheets_created.append("Sprint Issues")
        
        # Create Work Logs sheet
        if worklogs:
            ws_worklogs = self.wb.create_sheet(title="Work Logs")
            ws_worklogs.append([
                'Issue Key', 'Issue Type', 'Summary', 'Status', 
                'Author', 'Time Spent', 'Time Spent (Hours)', 'Date', 'Sprint', 'Comment'
            ])
            
            for log in worklogs:
                ws_worklogs.append([
                    log['issueKey'], log['issueType'], log['summary'], log['status'],
                    log['author'], log['timeSpent'], log['timeSpentHours'], 
                    log['startedDate'], log['sprint'], log['comment']
                ])
            
            sheets_created.append("Work Logs")
        
        # Create Comments sheet
        if comments:
            ws_comments = self.wb.create_sheet(title="Comments")
            ws_comments.append([
                'Issue Key', 'Summary', 'Status', 'Parent Summary', 
                'Issue Type', 'Comment Date', 'Comment Author', 'Comment'
            ])
            
            for comment in comments:
                ws_comments.append([
                    comment['issueKey'], comment['summary'], comment['status'],
                    comment['parent_summary'], comment['issueType'], comment['comment_date'],
                    comment['comment_author'], comment['comment_body']
                ])
            
            sheets_created.append("Comments")
        
        # Create Epics with Label sheet (conditional - only if epic_label_issues provided)
        if epic_label_issues:
            self._write_epic_sheet(epic_label_issues, "Epics with Label")
            sheets_created.append("Epics with Label")
        
        # Create Open Epics sheet (always created if open_epic_issues provided)
        if open_epic_issues:
            self._write_epic_sheet(open_epic_issues, "Open Epics")
            sheets_created.append("Open Epics")
        
        # Create Charts sheet if we have issues data
        if issues:
            self._create_charts_sheet(issues, worklogs, issues_by_sprint)
            sheets_created.append("Charts")
        
        # Create Progress sheet with progress charts
        if issues_by_sprint or epic_label_issues or open_epic_issues:
            progress_sheet = self._create_progress_sheet(issues_by_sprint, epic_label_issues, open_epic_issues)
            sheets_created.append(progress_sheet)
        
        # Create Time Tracking sheet with pivot tables (only if worklogs exist)
        if worklogs and len(worklogs) > 0:
            time_tracking_sheet = self._create_time_tracking_sheet()
            if time_tracking_sheet:
                sheets_created.append(time_tracking_sheet)
        
        if not sheets_created:
            return False, None, "No data was fetched to save."
        
        try:
            self.wb.save(filename)
            return True, filename, None
        except Exception as e:
            return False, None, str(e)
    
    def _create_charts_sheet(self, issues, worklogs=None, issues_by_sprint=None):
        """Creates a separate sheet for charts using the clean charts helper."""
        create_clean_charts_sheet(self.wb, issues, worklogs, issues_by_sprint)
    
    def _write_epic_sheet(self, epic_data, sheet_name):
        """
        Creates a sheet for epic-based issues (Epics with Label or Open Epics).
        
        Args:
            epic_data: Dictionary with 'issues' list and 'epic_statuses' dict
            sheet_name: Name of the sheet to create
        """
        ws = self.wb.create_sheet(title=sheet_name)
        ws.append(['Issue Key', 'Issue Type', 'Summary', 'Status', 'Sprint', 'Parent Summary', 
                   'Story Points', 'Parent Key', 'Status Category', 'Epic Status'])
        
        issues = epic_data.get('issues', [])
        epic_statuses = epic_data.get('epic_statuses', {})
        
        for issue in issues:
            parent_summary = 'N/A'
            parent_key = 'N/A'
            epic_status = 'N/A'
            
            parent_field = issue.get('fields', {}).get('parent')
            if parent_field:
                parent_summary = parent_field.get('fields', {}).get('summary', 'N/A')
                parent_key = parent_field.get('key', 'N/A')
                # Get epic status from the epic_statuses dict
                epic_status = epic_statuses.get(parent_key, 'N/A')
            
            # Get story points (use 0 if not set)
            story_points = issue.get('fields', {}).get(JIRA_STORY_POINTS_FIELD)
            if story_points is None:
                story_points = 0
            
            # Get status category
            status_category = issue.get('fields', {}).get('status', {}).get('statusCategory', {}).get('name', 'N/A')
            
            # Get sprint info if available (empty string if no sprint)
            sprint_name = issue.get('sprint_name', '')
            
            ws.append([
                issue.get('key'),
                issue.get('fields', {}).get('issuetype', {}).get('name', 'N/A'),
                issue.get('fields', {}).get('summary'),
                issue.get('fields', {}).get('status', {}).get('name', 'N/A'),
                sprint_name,
                parent_summary,
                story_points,
                parent_key,
                status_category,
                epic_status
            ])
    
    def _create_progress_sheet(self, issues_by_sprint=None, epic_label_issues=None, open_epic_issues=None):
        """
        Creates Progress sheet with progress visualization charts.
        
        Args:
            issues_by_sprint: Dict of sprint issues (sprint_id -> {name, issues})
            epic_label_issues: List of issues from "Epics with Label" sheet
            open_epic_issues: List of issues from "Open Epics" sheet
        """
        ws = self.wb.create_sheet(title="Progress")
        
        # Add sheet title
        ws['A1'] = "Progress Charts"
        ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)
        
        current_row = 3
        current_col = 1
        
        # Sprint Progress Charts (3 charts per sprint)
        if issues_by_sprint:
            for sprint_id, sprint_data in issues_by_sprint.items():
                sprint_issues = sprint_data['issues']
                sprint_name = sprint_data['name']
                
                # Calculate epic progress for this sprint
                epic_progress = calculate_epic_progress(sprint_issues)
                
                if epic_progress:  # Only create charts if there's data
                    # Chart 1: Sprint Progress in Percentage (Column A-B)
                    chart1, end_row1 = create_percentage_bar_chart(
                        ws, epic_progress, current_row,
                        f"{sprint_name} - Progress by Epic (%)"
                    )
                    ws.add_chart(chart1, f"A{current_row}")
                    
                    # Chart 2: Sprint Progress in Story Points (Stacked) (Column D-G)
                    chart2, end_row2 = create_stacked_bar_chart(
                        ws, epic_progress, current_row, 4,  # Start at column 4 (D)
                        f"{sprint_name} - Progress by Epic (Story Points)"
                    )
                    ws.add_chart(chart2, f"Q{current_row}")  # Column Q (17)
                    
                    # Chart 3: Sprint Composition (Pie) (Column I-J)
                    composition_data = calculate_sprint_composition(sprint_issues)
                    if composition_data:
                        chart3, end_row3 = create_composition_pie_chart(
                            ws, composition_data, current_row, 9,  # Start at column 9 (I)
                            f"{sprint_name} - Composition by Epic"
                        )
                        ws.add_chart(chart3, f"AG{current_row}")  # Column AG (33)
                    
                    # Move to next chart group
                    current_row = max(end_row1, end_row2, end_row3 if composition_data else end_row2) + 20
        
        # Epic Label Progress Charts (2 charts, conditional)
        if epic_label_issues and epic_label_issues.get('issues'):
            epic_progress = calculate_epic_progress(epic_label_issues['issues'])
            
            if epic_progress:
                # Chart 4: Epic Label Progress in Percentage (Column A-B)
                chart4, end_row4 = create_percentage_bar_chart(
                    ws, epic_progress, current_row,
                    "Epic Label Progress by Epic (%)"
                )
                ws.add_chart(chart4, f"A{current_row}")
                
                # Chart 5: Epic Label Progress in Story Points (Stacked) (Column D-G)
                chart5, end_row5 = create_stacked_bar_chart(
                    ws, epic_progress, current_row, 4,  # Start at column 4 (D)
                    "Epic Label Progress by Epic (Story Points)"
                )
                ws.add_chart(chart5, f"Q{current_row}")  # Column Q (17)
                
                # Move to next chart group
                current_row = max(end_row4, end_row5) + 20
        
        # Open Epic Progress Charts (2 charts, always created)
        if open_epic_issues and open_epic_issues.get('issues'):
            epic_progress = calculate_epic_progress(open_epic_issues['issues'])
            
            if epic_progress:
                # Chart 6: Open Epic Progress in Percentage (Column A-B)
                chart6, end_row6 = create_percentage_bar_chart(
                    ws, epic_progress, current_row,
                    "Open Epics Progress by Epic (%)"
                )
                ws.add_chart(chart6, f"A{current_row}")
                
                # Chart 7: Open Epic Progress in Story Points (Stacked) (Column D-G)
                chart7, end_row7 = create_stacked_bar_chart(
                    ws, epic_progress, current_row, 4,  # Start at column 4 (D)
                    "Open Epics Progress by Epic (Story Points)"
                )
                ws.add_chart(chart7, f"Q{current_row}")  # Column Q (17)
        
        return "Progress"
    
    def _create_time_tracking_sheet(self):
        """
        Creates Time Tracking sheet with aggregated time data for easy pivot table creation.
        
        Note: This creates formatted tables that users can easily convert to pivot tables in Excel.
        The data is pre-aggregated and organized for time tracking analysis.
        
        Returns:
            str: Sheet name if created, None otherwise
        """
        # Check if Work Logs sheet exists
        if "Work Logs" not in self.wb.sheetnames:
            return None
        
        work_logs_sheet = self.wb["Work Logs"]
        
        # Get worklog data from the sheet
        # Columns: Issue Key(A), Issue Type(B), Summary(C), Status(D), Author(E), Time Spent(F), Time Spent Hours(G), Date(H), Sprint(I), Comment(J)
        max_row = work_logs_sheet.max_row
        if max_row <= 1:  # Only header row or empty
            return None
        
        # Read all worklog data
        worklogs_data = []
        for row in range(2, max_row + 1):  # Skip header
            issue_key = work_logs_sheet.cell(row, 1).value
            author = work_logs_sheet.cell(row, 5).value
            date = work_logs_sheet.cell(row, 8).value  # Date is column H (8)
            hours = work_logs_sheet.cell(row, 7).value or 0  # Time Spent Hours is column G (7)
            
            worklogs_data.append({
                'issue_key': issue_key,
                'author': author,
                'date': date,
                'hours': float(hours) if hours else 0
            })
        
        # Create Time Tracking sheet
        ws = self.wb.create_sheet(title="Time Tracking")
        
        # Add title and instructions
        ws['A1'] = "Time Tracking Analysis"
        ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)
        ws['A2'] = "Tip: Select any table below and use Insert > PivotTable in Excel for interactive analysis"
        ws['A2'].font = ws['A2'].font.copy(italic=True, size=10)
        
        # Table 1: Time by Date, Author, and Issue
        # This provides the detailed view requested
        ws['A4'] = "Detailed Time Tracking"
        ws['A4'].font = ws['A4'].font.copy(bold=True, size=12)
        
        # Headers for Table 1
        current_row = 6
        ws.cell(current_row, 1, "Date")
        ws.cell(current_row, 2, "Author")
        ws.cell(current_row, 3, "Issue Key")
        ws.cell(current_row, 4, "Hours")
        
        # Make headers bold
        for col in range(1, 5):
            ws.cell(current_row, col).font = ws.cell(current_row, col).font.copy(bold=True)
        
        # Sort worklogs by date, then author, then issue
        sorted_worklogs = sorted(worklogs_data, key=lambda x: (x['date'] or '', x['author'] or '', x['issue_key'] or ''))
        
        # Write data
        current_row += 1
        start_row_table1 = current_row
        for wl in sorted_worklogs:
            ws.cell(current_row, 1, wl['date'])
            ws.cell(current_row, 2, wl['author'])
            ws.cell(current_row, 3, wl['issue_key'])
            ws.cell(current_row, 4, wl['hours'])
            current_row += 1
        
        end_row_table1 = current_row - 1
        
        # Create Excel Table for easy filtering and pivot table creation
        if end_row_table1 >= start_row_table1:
            table1 = Table(displayName="DetailedTimeTracking", ref=f"A6:D{end_row_table1}")
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table1.tableStyleInfo = style
            ws.add_table(table1)
        
        # Table 2: Summary by Author and Date
        # This provides aggregated view
        current_row += 3
        ws.cell(current_row, 1, "Summary by Author and Date")
        ws.cell(current_row, 1).font = ws.cell(current_row, 1).font.copy(bold=True, size=12)
        
        current_row += 2
        ws.cell(current_row, 1, "Author")
        ws.cell(current_row, 2, "Date")
        ws.cell(current_row, 3, "Total Hours")
        
        # Make headers bold
        for col in range(1, 4):
            ws.cell(current_row, col).font = ws.cell(current_row, col).font.copy(bold=True)
        
        # Aggregate by author and date
        from collections import defaultdict
        author_date_hours = defaultdict(float)
        for wl in worklogs_data:
            key = (wl['author'], wl['date'])
            author_date_hours[key] += wl['hours']
        
        # Sort by author, then date
        sorted_summary = sorted(author_date_hours.items(), key=lambda x: (x[0][0] or '', x[0][1] or ''))
        
        current_row += 1
        start_row_table2 = current_row
        for (author, date), hours in sorted_summary:
            ws.cell(current_row, 1, author)
            ws.cell(current_row, 2, date)
            ws.cell(current_row, 3, hours)
            current_row += 1
        
        end_row_table2 = current_row - 1
        
        # Create Excel Table
        if end_row_table2 >= start_row_table2:
            table2 = Table(displayName="SummaryByAuthorDate", ref=f"A{start_row_table2-1}:C{end_row_table2}")
            style2 = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table2.tableStyleInfo = style2
            ws.add_table(table2)
        
        # Table 3: Summary by Author (Total)
        current_row += 3
        ws.cell(current_row, 1, "Total Hours by Author")
        ws.cell(current_row, 1).font = ws.cell(current_row, 1).font.copy(bold=True, size=12)
        
        current_row += 2
        ws.cell(current_row, 1, "Author")
        ws.cell(current_row, 2, "Total Hours")
        
        # Make headers bold
        for col in range(1, 3):
            ws.cell(current_row, col).font = ws.cell(current_row, col).font.copy(bold=True)
        
        # Aggregate by author only
        author_hours = defaultdict(float)
        for wl in worklogs_data:
            author_hours[wl['author']] += wl['hours']
        
        # Sort by author
        sorted_authors = sorted(author_hours.items(), key=lambda x: x[0] or '')
        
        current_row += 1
        start_row_table3 = current_row
        for author, hours in sorted_authors:
            ws.cell(current_row, 1, author)
            ws.cell(current_row, 2, hours)
            current_row += 1
        
        end_row_table3 = current_row - 1
        
        # Create Excel Table
        if end_row_table3 >= start_row_table3:
            table3 = Table(displayName="TotalByAuthor", ref=f"A{start_row_table3-1}:B{end_row_table3}")
            style3 = TableStyleInfo(
                name="TableStyleMedium6",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table3.tableStyleInfo = style3
            ws.add_table(table3)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        
        return "Time Tracking"
    
    def get_workbook(self):
        """Returns the current workbook instance."""
        return self.wb
