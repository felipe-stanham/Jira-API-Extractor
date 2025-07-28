"""Improved helper functions for creating charts in Excel with better formatting."""

from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.colors import ColorChoice
from collections import Counter
from chart_colors import assign_colors_to_series, get_issue_type_color, get_status_color

def apply_colors_to_pie_chart(pie_chart, items, color_map_func):
    """Apply consistent colors to pie chart series based on configuration."""
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.fill import PatternFillProperties
    from openpyxl.drawing.colors import ColorChoice
    
    color_assignments = assign_colors_to_series(items, color_map_func)
    
    # Create data points for each slice with colors
    data_points = []
    for i, item in enumerate(items):
        color_hex = color_assignments[item]
        
        # Create data point with solid fill color
        dp = DataPoint(idx=i)
        dp.spPr = GraphicalProperties()
        dp.spPr.solidFill = ColorChoice(srgbClr=color_hex)
        data_points.append(dp)
    
    # Apply data points to the first series
    if len(pie_chart.series) > 0:
        pie_chart.series[0].data_points = data_points

def create_clean_charts_sheet(wb, issues, worklogs=None, issues_by_sprint=None):
    """Creates a charts sheet with improved formatting and labels."""
    ws_charts = wb.create_sheet(title="Charts")
    
    if not issues:
        ws_charts.append(['No data available for charts'])
        return
    
    # ===== ISSUES BY STATUS CHART =====
    ws_charts.append(['Issues by Status Analysis'])
    ws_charts.append(['Status', 'Count'])
    
    # Count issues by status using Python instead of Excel formulas
    status_counts = Counter()
    for issue in issues:
        status = issue.get('fields', {}).get('status', {}).get('name', 'Unknown')
        status_counts[status] += 1
    
    status_start_row = 3
    for i, (status, count) in enumerate(sorted(status_counts.items())):
        row = status_start_row + i
        ws_charts.cell(row=row, column=1, value=status)
        ws_charts.cell(row=row, column=2, value=count)
    
    status_end_row = status_start_row + len(status_counts) - 1
    
    # Create pie chart for Issues by Status
    pie_status = PieChart()
    pie_status.title = "Issues by Status"
    pie_status.width = 12
    pie_status.height = 8
    
    labels = Reference(ws_charts, min_col=1, min_row=status_start_row, max_row=status_end_row)
    data = Reference(ws_charts, min_col=2, min_row=status_start_row, max_row=status_end_row)
    
    pie_status.add_data(data, titles_from_data=False)
    pie_status.set_categories(labels)
    
    # Apply colors based on status configuration
    apply_colors_to_pie_chart(pie_status, sorted(status_counts.keys()), get_status_color)
    
    # Configure chart appearance - show only value and percentage
    pie_status.dataLabels = DataLabelList()
    pie_status.dataLabels.showCatName = False  # Don't show category name
    pie_status.dataLabels.showVal = True       # Show value
    pie_status.dataLabels.showPercent = True   # Show percentage
    pie_status.dataLabels.showSerName = False  # Don't show series name
    
    ws_charts.add_chart(pie_status, "D2")
    
    # ===== ISSUES BY TYPE CHART =====
    type_start_row = status_end_row + 3
    ws_charts.cell(row=type_start_row - 1, column=1, value='Issues by Type Analysis')
    ws_charts.cell(row=type_start_row, column=1, value='Issue Type')
    ws_charts.cell(row=type_start_row, column=2, value='Count')
    
    # Count issues by type using Python
    type_counts = Counter()
    for issue in issues:
        issue_type = issue.get('fields', {}).get('issuetype', {}).get('name', 'Unknown')
        type_counts[issue_type] += 1
    
    type_data_start = type_start_row + 1
    for i, (issue_type, count) in enumerate(sorted(type_counts.items())):
        row = type_data_start + i
        ws_charts.cell(row=row, column=1, value=issue_type)
        ws_charts.cell(row=row, column=2, value=count)
    
    type_end_row = type_data_start + len(type_counts) - 1
    
    # Create pie chart for Issues by Type
    pie_type = PieChart()
    pie_type.title = "Issues by Type"
    pie_type.width = 12
    pie_type.height = 8
    
    labels_type = Reference(ws_charts, min_col=1, min_row=type_data_start, max_row=type_end_row)
    data_type = Reference(ws_charts, min_col=2, min_row=type_data_start, max_row=type_end_row)
    
    pie_type.add_data(data_type, titles_from_data=False)
    pie_type.set_categories(labels_type)
    
    # Apply colors based on issue type configuration
    apply_colors_to_pie_chart(pie_type, sorted(type_counts.keys()), get_issue_type_color)
    
    # Configure chart appearance - show only value and percentage
    pie_type.dataLabels = DataLabelList()
    pie_type.dataLabels.showCatName = False  # Don't show category name
    pie_type.dataLabels.showVal = True       # Show value
    pie_type.dataLabels.showPercent = True   # Show percentage
    pie_type.dataLabels.showSerName = False  # Don't show series name
    
    ws_charts.add_chart(pie_type, "D15")
    
    # ===== TIME BY ISSUE TYPE CHART =====
    if worklogs:
        time_start_row = type_end_row + 3
        ws_charts.cell(row=time_start_row - 1, column=1, value='Total Time by Issue Type')
        ws_charts.cell(row=time_start_row, column=1, value='Issue Type')
        ws_charts.cell(row=time_start_row, column=2, value='Total Hours')
        
        # Calculate time by issue type using Python
        time_by_type = {}
        for log in worklogs:
            issue_type = log.get('issueType', 'Unknown')
            hours = log.get('timeSpentHours', 0)
            if issue_type in time_by_type:
                time_by_type[issue_type] += hours
            else:
                time_by_type[issue_type] = hours
        
        time_data_start = time_start_row + 1
        for i, (issue_type, hours) in enumerate(sorted(time_by_type.items())):
            row = time_data_start + i
            ws_charts.cell(row=row, column=1, value=issue_type)
            ws_charts.cell(row=row, column=2, value=round(hours, 2))
        
        time_end_row = time_data_start + len(time_by_type) - 1
        
        # Create pie chart for Time by Issue Type
        pie_time = PieChart()
        pie_time.title = "Total Time by Issue Type"
        pie_time.width = 12
        pie_time.height = 8
        
        labels_time = Reference(ws_charts, min_col=1, min_row=time_data_start, max_row=time_end_row)
        data_time = Reference(ws_charts, min_col=2, min_row=time_data_start, max_row=time_end_row)
        
        pie_time.add_data(data_time, titles_from_data=False)
        pie_time.set_categories(labels_time)
        
        # Apply colors based on issue type configuration
        apply_colors_to_pie_chart(pie_time, sorted(time_by_type.keys()), get_issue_type_color)
        
        # Configure chart appearance - show only value and percentage
        pie_time.dataLabels = DataLabelList()
        pie_time.dataLabels.showCatName = False  # Don't show category name
        pie_time.dataLabels.showVal = True       # Show value (hours)
        pie_time.dataLabels.showPercent = True   # Show percentage
        pie_time.dataLabels.showSerName = False  # Don't show series name
        
        ws_charts.add_chart(pie_time, "D28")
        
        # ===== IMPROVED TIME BY AUTHOR AND ISSUE TYPE BAR CHART =====
        author_start_row = time_end_row + 3
        ws_charts.cell(row=author_start_row - 1, column=1, value='Time Spent by Author and Issue Type')
        
        # Calculate time by author and issue type using Python
        author_type_time = {}
        author_totals = {}
        unique_authors = set()
        unique_types = set()
        
        for log in worklogs:
            author = log.get('author', 'Unknown')
            issue_type = log.get('issueType', 'Unknown')
            hours = log.get('timeSpentHours', 0)
            
            unique_authors.add(author)
            unique_types.add(issue_type)
            
            # Track time by author and type
            key = (author, issue_type)
            if key in author_type_time:
                author_type_time[key] += hours
            else:
                author_type_time[key] = hours
            
            # Track total by author
            if author in author_totals:
                author_totals[author] += hours
            else:
                author_totals[author] = hours
        
        unique_authors = sorted(unique_authors)
        unique_types = sorted(unique_types)
        
        # Create header row: Author, Issue Type 1, Issue Type 2, ..., Total
        header_row = author_start_row
        ws_charts.cell(row=header_row, column=1, value='Author')
        
        col = 2
        for issue_type in unique_types:
            ws_charts.cell(row=header_row, column=col, value=issue_type)
            col += 1
        ws_charts.cell(row=header_row, column=col, value='Total')
        
        # Create data rows
        data_start_row = author_start_row + 1
        for i, author in enumerate(unique_authors):
            row = data_start_row + i
            ws_charts.cell(row=row, column=1, value=author)
            
            col = 2
            for issue_type in unique_types:
                hours = author_type_time.get((author, issue_type), 0)
                ws_charts.cell(row=row, column=col, value=round(hours, 2))
                col += 1
            
            # Total for author
            total_hours = author_totals.get(author, 0)
            ws_charts.cell(row=row, column=col, value=round(total_hours, 2))
        
        author_end_row = data_start_row + len(unique_authors) - 1
        
        # Create improved bar chart
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "Time Spent by Author and Issue Type"
        bar_chart.y_axis.title = 'Hours'
        bar_chart.x_axis.title = 'Authors'
        bar_chart.width = 20
        bar_chart.height = 12
        
        # Add data series for each issue type (not including total column)
        for i, issue_type in enumerate(unique_types):
            data_col = 2 + i
            data_ref = Reference(ws_charts, min_col=data_col, min_row=data_start_row, max_row=author_end_row)
            bar_chart.add_data(data_ref, titles_from_data=False)
        
        # Add total series
        total_col = 2 + len(unique_types)
        total_data_ref = Reference(ws_charts, min_col=total_col, min_row=data_start_row, max_row=author_end_row)
        bar_chart.add_data(total_data_ref, titles_from_data=False)
        
        # Set series titles manually after adding data
        for i, issue_type in enumerate(unique_types):
            if i < len(bar_chart.series):
                series_label = SeriesLabel()
                series_label.v = issue_type
                bar_chart.series[i].tx = series_label
        
        # Set title for total series
        if len(bar_chart.series) > len(unique_types):
            total_label = SeriesLabel()
            total_label.v = "Total"
            bar_chart.series[len(unique_types)].tx = total_label
        
        # Set categories (authors)
        cats = Reference(ws_charts, min_col=1, min_row=data_start_row, max_row=author_end_row)
        bar_chart.set_categories(cats)
        
        # Enable data labels to show hours on each bar
        bar_chart.dataLabels = DataLabelList()
        bar_chart.dataLabels.showVal = True
        
        ws_charts.add_chart(bar_chart, "D41")
        
        # Add summary information
        summary_row = author_end_row + 3
        ws_charts.cell(row=summary_row, column=1, value='Summary Information')
        ws_charts.cell(row=summary_row + 1, column=1, value='Total Issues:')
        ws_charts.cell(row=summary_row + 1, column=2, value=len(issues))
        
        ws_charts.cell(row=summary_row + 2, column=1, value='Total Work Log Entries:')
        ws_charts.cell(row=summary_row + 2, column=2, value=len(worklogs))
        
        total_hours = sum(log.get('timeSpentHours', 0) for log in worklogs)
        ws_charts.cell(row=summary_row + 3, column=1, value='Total Hours Logged:')
        ws_charts.cell(row=summary_row + 3, column=2, value=round(total_hours, 2))
        
        # Add individual sprint charts if multiple sprints provided
        if issues_by_sprint and len(issues_by_sprint) > 1:
            current_row = summary_row + 5
            chart_col = "P"  # Column P for sprint-specific charts
            
            for sprint_id, sprint_data in issues_by_sprint.items():
                sprint_issues = sprint_data['issues']
                sprint_name = sprint_data['name']
                
                # ===== SPRINT STATUS CHART =====
                # Create data for this sprint's status chart
                ws_charts.cell(row=current_row, column=1, value=f'{sprint_name} - Issues by Status')
                ws_charts.cell(row=current_row + 1, column=1, value='Status')
                ws_charts.cell(row=current_row + 1, column=2, value='Count')
                
                # Count issues by status for this sprint
                sprint_status_counts = Counter()
                for issue in sprint_issues:
                    status = issue.get('fields', {}).get('status', {}).get('name', 'Unknown')
                    sprint_status_counts[status] += 1
                
                sprint_status_start = current_row + 2
                for i, (status, count) in enumerate(sorted(sprint_status_counts.items())):
                    row = sprint_status_start + i
                    ws_charts.cell(row=row, column=1, value=status)
                    ws_charts.cell(row=row, column=2, value=count)
                
                sprint_status_end = sprint_status_start + len(sprint_status_counts) - 1
                
                # Create pie chart for this sprint's status
                pie_sprint_status = PieChart()
                pie_sprint_status.title = f"{sprint_name} - Issues by Status"
                pie_sprint_status.width = 10
                pie_sprint_status.height = 7
                
                labels_sprint_status = Reference(ws_charts, min_col=1, min_row=sprint_status_start, max_row=sprint_status_end)
                data_sprint_status = Reference(ws_charts, min_col=2, min_row=sprint_status_start, max_row=sprint_status_end)
                
                pie_sprint_status.add_data(data_sprint_status, titles_from_data=False)
                pie_sprint_status.set_categories(labels_sprint_status)
                
                # Apply colors based on status configuration
                apply_colors_to_pie_chart(pie_sprint_status, sorted(sprint_status_counts.keys()), get_status_color)
                
                # Configure chart appearance - show only value and percentage
                pie_sprint_status.dataLabels = DataLabelList()
                pie_sprint_status.dataLabels.showCatName = False  # Don't show category name
                pie_sprint_status.dataLabels.showVal = True       # Show value
                pie_sprint_status.dataLabels.showPercent = True   # Show percentage
                pie_sprint_status.dataLabels.showSerName = False  # Don't show series name
                
                # Position status chart
                chart_position_status = f"{chart_col}{current_row}"
                ws_charts.add_chart(pie_sprint_status, chart_position_status)
                
                # ===== SPRINT TYPE CHART =====
                # Move to next section for type chart
                current_row = sprint_status_end + 3
                
                # Create data for this sprint's type chart
                ws_charts.cell(row=current_row, column=1, value=f'{sprint_name} - Issues by Type')
                ws_charts.cell(row=current_row + 1, column=1, value='Issue Type')
                ws_charts.cell(row=current_row + 1, column=2, value='Count')
                
                # Count issues by type for this sprint
                sprint_type_counts = Counter()
                for issue in sprint_issues:
                    issue_type = issue.get('fields', {}).get('issuetype', {}).get('name', 'Unknown')
                    sprint_type_counts[issue_type] += 1
                
                sprint_type_start = current_row + 2
                for i, (issue_type, count) in enumerate(sorted(sprint_type_counts.items())):
                    row = sprint_type_start + i
                    ws_charts.cell(row=row, column=1, value=issue_type)
                    ws_charts.cell(row=row, column=2, value=count)
                
                sprint_type_end = sprint_type_start + len(sprint_type_counts) - 1
                
                # Create pie chart for this sprint's type
                pie_sprint_type = PieChart()
                pie_sprint_type.title = f"{sprint_name} - Issues by Type"
                pie_sprint_type.width = 10
                pie_sprint_type.height = 7
                
                labels_sprint_type = Reference(ws_charts, min_col=1, min_row=sprint_type_start, max_row=sprint_type_end)
                data_sprint_type = Reference(ws_charts, min_col=2, min_row=sprint_type_start, max_row=sprint_type_end)
                
                pie_sprint_type.add_data(data_sprint_type, titles_from_data=False)
                pie_sprint_type.set_categories(labels_sprint_type)
                
                # Apply colors based on issue type configuration
                apply_colors_to_pie_chart(pie_sprint_type, sorted(sprint_type_counts.keys()), get_issue_type_color)
                
                # Configure chart appearance - show only value and percentage
                pie_sprint_type.dataLabels = DataLabelList()
                pie_sprint_type.dataLabels.showCatName = False  # Don't show category name
                pie_sprint_type.dataLabels.showVal = True       # Show value
                pie_sprint_type.dataLabels.showPercent = True   # Show percentage
                pie_sprint_type.dataLabels.showSerName = False  # Don't show series name
                
                # Position type chart (next to status chart)
                chart_position_type = f"AB{current_row}"  # Column AB for type charts
                ws_charts.add_chart(pie_sprint_type, chart_position_type)
                
                # Move to next sprint position
                current_row = sprint_type_end + 5
    else:
        # Add summary for issues only
        summary_row = type_end_row + 3
        ws_charts.cell(row=summary_row, column=1, value='Summary Information')
        ws_charts.cell(row=summary_row + 1, column=1, value='Total Issues:')
        ws_charts.cell(row=summary_row + 1, column=2, value=len(issues))
        
        # Add individual sprint charts if multiple sprints provided
        if issues_by_sprint and len(issues_by_sprint) > 1:
            current_row = summary_row + 3
            chart_col = "P"  # Column P for sprint-specific charts
            
            for sprint_id, sprint_data in issues_by_sprint.items():
                sprint_issues = sprint_data['issues']
                sprint_name = sprint_data['name']
                
                # ===== SPRINT STATUS CHART =====
                # Create data for this sprint's status chart
                ws_charts.cell(row=current_row, column=1, value=f'{sprint_name} - Issues by Status')
                ws_charts.cell(row=current_row + 1, column=1, value='Status')
                ws_charts.cell(row=current_row + 1, column=2, value='Count')
                
                # Count issues by status for this sprint
                sprint_status_counts = Counter()
                for issue in sprint_issues:
                    status = issue.get('fields', {}).get('status', {}).get('name', 'Unknown')
                    sprint_status_counts[status] += 1
                
                sprint_status_start = current_row + 2
                for i, (status, count) in enumerate(sorted(sprint_status_counts.items())):
                    row = sprint_status_start + i
                    ws_charts.cell(row=row, column=1, value=status)
                    ws_charts.cell(row=row, column=2, value=count)
                
                sprint_status_end = sprint_status_start + len(sprint_status_counts) - 1
                
                # Create pie chart for this sprint's status
                pie_sprint_status = PieChart()
                pie_sprint_status.title = f"{sprint_name} - Issues by Status"
                pie_sprint_status.width = 10
                pie_sprint_status.height = 7
                
                labels_sprint_status = Reference(ws_charts, min_col=1, min_row=sprint_status_start, max_row=sprint_status_end)
                data_sprint_status = Reference(ws_charts, min_col=2, min_row=sprint_status_start, max_row=sprint_status_end)
                
                pie_sprint_status.add_data(data_sprint_status, titles_from_data=False)
                pie_sprint_status.set_categories(labels_sprint_status)
                
                # Apply colors based on status configuration
                apply_colors_to_pie_chart(pie_sprint_status, sorted(sprint_status_counts.keys()), get_status_color)
                
                # Configure chart appearance - show only value and percentage
                pie_sprint_status.dataLabels = DataLabelList()
                pie_sprint_status.dataLabels.showCatName = False  # Don't show category name
                pie_sprint_status.dataLabels.showVal = True       # Show value
                pie_sprint_status.dataLabels.showPercent = True   # Show percentage
                pie_sprint_status.dataLabels.showSerName = False  # Don't show series name
                
                # Position status chart
                chart_position_status = f"{chart_col}{current_row}"
                ws_charts.add_chart(pie_sprint_status, chart_position_status)
                
                # ===== SPRINT TYPE CHART =====
                # Move to next section for type chart
                current_row = sprint_status_end + 3
                
                # Create data for this sprint's type chart
                ws_charts.cell(row=current_row, column=1, value=f'{sprint_name} - Issues by Type')
                ws_charts.cell(row=current_row + 1, column=1, value='Issue Type')
                ws_charts.cell(row=current_row + 1, column=2, value='Count')
                
                # Count issues by type for this sprint
                sprint_type_counts = Counter()
                for issue in sprint_issues:
                    issue_type = issue.get('fields', {}).get('issuetype', {}).get('name', 'Unknown')
                    sprint_type_counts[issue_type] += 1
                
                sprint_type_start = current_row + 2
                for i, (issue_type, count) in enumerate(sorted(sprint_type_counts.items())):
                    row = sprint_type_start + i
                    ws_charts.cell(row=row, column=1, value=issue_type)
                    ws_charts.cell(row=row, column=2, value=count)
                
                sprint_type_end = sprint_type_start + len(sprint_type_counts) - 1
                
                # Create pie chart for this sprint's type
                pie_sprint_type = PieChart()
                pie_sprint_type.title = f"{sprint_name} - Issues by Type"
                pie_sprint_type.width = 10
                pie_sprint_type.height = 7
                
                labels_sprint_type = Reference(ws_charts, min_col=1, min_row=sprint_type_start, max_row=sprint_type_end)
                data_sprint_type = Reference(ws_charts, min_col=2, min_row=sprint_type_start, max_row=sprint_type_end)
                
                pie_sprint_type.add_data(data_sprint_type, titles_from_data=False)
                pie_sprint_type.set_categories(labels_sprint_type)
                
                # Apply colors based on issue type configuration
                apply_colors_to_pie_chart(pie_sprint_type, sorted(sprint_type_counts.keys()), get_issue_type_color)
                
                # Configure chart appearance - show only value and percentage
                pie_sprint_type.dataLabels = DataLabelList()
                pie_sprint_type.dataLabels.showCatName = False  # Don't show category name
                pie_sprint_type.dataLabels.showVal = True       # Show value
                pie_sprint_type.dataLabels.showPercent = True   # Show percentage
                pie_sprint_type.dataLabels.showSerName = False  # Don't show series name
                
                # Position type chart (next to status chart)
                chart_position_type = f"AB{current_row}"  # Column AB for type charts
                ws_charts.add_chart(pie_sprint_type, chart_position_type)
                
                # Move to next sprint position
                current_row = sprint_type_end + 5
