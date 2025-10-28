#!/usr/bin/env python3
"""
Progress Charts Helper Module

Creates progress visualization charts for the Progress sheet.
Includes horizontal bar charts (percentage and stacked story points)
and pie charts (sprint composition).
"""

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import SolidColorFillProperties, ColorChoice


# Progress state colors
PROGRESS_COLORS = {
    'Done': '4CAF50',      # Green
    'In Progress': 'FFC107',  # Yellow
    'To Do': '2196F3'      # Blue
}


def create_percentage_bar_chart(ws, epic_progress_data, start_row, chart_title):
    """
    Creates horizontal bar chart showing epic progress by percentage.
    
    Args:
        ws: Worksheet object
        epic_progress_data: List of dicts from calculate_epic_progress()
        start_row: Starting row for data
        chart_title: Title for the chart
        
    Returns:
        BarChart object
    """
    # Write data to worksheet
    ws.cell(row=start_row, column=1, value="Epic")
    ws.cell(row=start_row, column=2, value="Completion %")
    
    current_row = start_row + 1
    for epic in epic_progress_data:
        ws.cell(row=current_row, column=1, value=epic['epic_name'])
        ws.cell(row=current_row, column=2, value=round(epic['percentage'], 1))
        current_row += 1
    
    # Create bar chart
    chart = BarChart()
    chart.type = "bar"  # Horizontal bars
    chart.title = chart_title
    chart.x_axis.title = "Completion %"
    chart.y_axis.title = "Epic"
    
    # Add data
    data = Reference(ws, min_col=2, min_row=start_row, max_row=current_row - 1)
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=current_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Style
    chart.height = max(10, len(epic_progress_data) * 0.8)  # Dynamic height
    chart.width = 15
    chart.legend = None  # No legend needed for single series
    
    return chart, current_row


def create_stacked_bar_chart(ws, epic_progress_data, start_row, start_col, chart_title):
    """
    Creates horizontal stacked bar chart showing epic progress by story points.
    Stacked in order: Done (left) → In Progress (middle) → To Do (right)
    
    Args:
        ws: Worksheet object
        epic_progress_data: List of dicts from calculate_epic_progress()
        start_row: Starting row for data
        start_col: Starting column for data
        chart_title: Title for the chart
        
    Returns:
        BarChart object
    """
    # Write headers
    ws.cell(row=start_row, column=start_col, value="Epic")
    ws.cell(row=start_row, column=start_col + 1, value="Done")
    ws.cell(row=start_row, column=start_col + 2, value="In Progress")
    ws.cell(row=start_row, column=start_col + 3, value="To Do")
    
    # Write data
    current_row = start_row + 1
    for epic in epic_progress_data:
        ws.cell(row=current_row, column=start_col, value=epic['epic_name'])
        ws.cell(row=current_row, column=start_col + 1, value=epic['done_points'])
        ws.cell(row=current_row, column=start_col + 2, value=epic['in_progress_points'])
        ws.cell(row=current_row, column=start_col + 3, value=epic['to_do_points'])
        current_row += 1
    
    # Create stacked bar chart
    chart = BarChart()
    chart.type = "bar"  # Horizontal bars
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = chart_title
    chart.x_axis.title = "Story Points"
    chart.y_axis.title = "Epic"
    
    # Add data series (Done, In Progress, To Do)
    data = Reference(ws, min_col=start_col + 1, max_col=start_col + 3, min_row=start_row, max_row=current_row - 1)
    cats = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=current_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Apply colors to series
    if len(chart.series) >= 3:
        # Done - Green
        chart.series[0].graphicalProperties.solidFill = PROGRESS_COLORS['Done']
        # In Progress - Yellow
        chart.series[1].graphicalProperties.solidFill = PROGRESS_COLORS['In Progress']
        # To Do - Blue
        chart.series[2].graphicalProperties.solidFill = PROGRESS_COLORS['To Do']
    
    # Style
    chart.height = max(10, len(epic_progress_data) * 0.8)  # Dynamic height
    chart.width = 15
    chart.legend.position = 'b'  # Bottom legend
    
    # Enable data labels
    for series in chart.series:
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
    
    return chart, current_row


def create_composition_pie_chart(ws, composition_data, start_row, start_col, chart_title):
    """
    Creates pie chart showing sprint composition by epic.
    Shows total story points per epic (all statuses).
    
    Args:
        ws: Worksheet object
        composition_data: List of dicts from calculate_sprint_composition()
        start_row: Starting row for data
        start_col: Starting column for data
        chart_title: Title for the chart
        
    Returns:
        PieChart object
    """
    # Write data to worksheet
    ws.cell(row=start_row, column=start_col, value="Epic")
    ws.cell(row=start_row, column=start_col + 1, value="Story Points")
    
    current_row = start_row + 1
    for epic in composition_data:
        ws.cell(row=current_row, column=start_col, value=epic['epic_name'])
        ws.cell(row=current_row, column=start_col + 1, value=epic['total_points'])
        current_row += 1
    
    # Create pie chart
    chart = PieChart()
    chart.title = chart_title
    
    # Add data
    data = Reference(ws, min_col=start_col + 1, min_row=start_row, max_row=current_row - 1)
    cats = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=current_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Configure data labels
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showCatName = False
    chart.dataLabels.showVal = True
    chart.dataLabels.showPercent = True
    chart.dataLabels.showSerName = False
    
    # Style
    chart.height = 12
    chart.width = 15
    
    return chart, current_row


# Import DataLabelList for data labels
try:
    from openpyxl.chart.label import DataLabelList
except ImportError:
    # Fallback for older openpyxl versions
    class DataLabelList:
        def __init__(self):
            self.showCatName = False
            self.showVal = True
            self.showPercent = False
            self.showSerName = False
