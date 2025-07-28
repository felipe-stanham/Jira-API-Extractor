"""Enhanced helper functions for creating charts in Excel with better formatting and new chart types."""

from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.colors import ColorChoice
from collections import Counter, defaultdict
from chart_colors import assign_colors_to_series, get_issue_type_color, get_status_color

def apply_colors_to_pie_chart(pie_chart, items, color_map_func):
    """Apply consistent colors to pie chart series based on configuration."""
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.fill import PatternFillProperties
    from openpyxl.drawing.colors import ColorChoice
    
    color_assignments = assign_colors_to_series(items, color_map_func)
    
    # Create data points for each slice with colors
    data_points = []
    for i, item in enumerate(items):
        color_hex = color_assignments[item]
        
        # Create data point with solid fill color
        dp = DataPoint(idx=i)
        dp.spPr = GraphicalProperties()
        dp.spPr.solidFill = ColorChoice(srgbClr=color_hex)
        data_points.append(dp)
    
    # Apply data points to the first series
    if len(pie_chart.series) > 0:
        pie_chart.series[0].data_points = data_points

def create_clean_charts_sheet(wb, issues, worklogs=None, issues_by_sprint=None):
    """Creates a charts sheet with improved formatting and labels."""
    ws_charts = wb.create_sheet(title="Charts")
    
    if not issues:
        ws_charts.append(['No data available for charts'])
        return
    
    # ===== ISSUES BY STATUS CHART =====
    ws_charts.append(['Issues by Status Analysis'])
    ws_charts.append(['Status', 'Count'])
    
    # Count issues by status using Python instead of Excel formulas
    status_counts = Counter()
    for issue in issues:
        status = issue.get('fields', {}).get('status', {}).get('name', 'Unknown')
        status_counts[status] += 1
    
    status_start_row = 3
    for i, (status, count) in enumerate(sorted(status_counts.items())):
        row = status_start_row + i
        ws_charts.cell(row=row, column=1, value=status)
        ws_charts.cell(row=row, column=2, value=count)
    
    status_end_row = status_start_row + len(status_counts) - 1
    
    # Create pie chart for Issues by Status
    pie_status = PieChart()
    pie_status.title = "Issues by Status"
    pie_status.width = 12
    pie_status.height = 8
    
    labels = Reference(ws_charts, min_col=1, min_row=status_start_row, max_row=status_end_row)
    data = Reference(ws_charts, min_col=2, min_row=status_start_row, max_row=status_end_row)
    
    pie_status.add_data(data, titles_from_data=False)
    pie_status.set_categories(labels)
    
    # Apply colors based on status configuration
    apply_colors_to_pie_chart(pie_status, sorted(status_counts.keys()), get_status_color)
    
    # Configure chart appearance - show only value and percentage
    pie_status.dataLabels = DataLabelList()
    pie_status.dataLabels.showCatName = False  # Don't show category name
    pie_status.dataLabels.showVal = True       # Show value
    pie_status.dataLabels.showPercent = True   # Show percentage
    pie_status.dataLabels.showSerName = False  # Don't show series name
    
    ws_charts.add_chart(pie_status, "D2")
    
    # ===== ISSUES BY TYPE CHART =====
    type_start_row = status_end_row + 3
    ws_charts.cell(row=type_start_row - 1, column=1, value='Issues by Type Analysis')
    ws_charts.cell(row=type_start_row, column=1, value='Issue Type')
    ws_charts.cell(row=type_start_row, column=2, value='Count')
    
    # Count issues by type using Python
    type_counts = Counter()
    for issue in issues:
        issue_type = issue.get('fields', {}).get('issuetype', {}).get('name', 'Unknown')
        type_counts[issue_type] += 1
    
    type_data_start = type_start_row + 1
    for i, (issue_type, count) in enumerate(sorted(type_counts.items())):
        row = type_data_start + i
        ws_charts.cell(row=row, column=1, value=issue_type)
        ws_charts.cell(row=row, column=2, value=count)
    
    type_end_row = type_data_start + len(type_counts) - 1
    
    # Create pie chart for Issues by Type
    pie_type = PieChart()
    pie_type.title = "Issues by Type"
    pie_type.width = 12
    pie_type.height = 8
    
    labels_type = Reference(ws_charts, min_col=1, min_row=type_data_start, max_row=type_end_row)
    data_type = Reference(ws_charts, min_col=2, min_row=type_data_start, max_row=type_end_row)
    
    pie_type.add_data(data_type, titles_from_data=False)
    pie_type.set_categories(labels_type)
    
    # Apply colors based on issue type configuration
    apply_colors_to_pie_chart(pie_type, sorted(type_counts.keys()), get_issue_type_color)
    
    # Configure chart appearance - show only value and percentage
    pie_type.dataLabels = DataLabelList()
    pie_type.dataLabels.showCatName = False  # Don't show category name
    pie_type.dataLabels.showVal = True       # Show value
    pie_type.dataLabels.showPercent = True   # Show percentage
    pie_type.dataLabels.showSerName = False  # Don't show series name
    
    ws_charts.add_chart(pie_type, "P2")
    
    # ===== TIME-BASED CHARTS (if worklogs available) =====
    if worklogs:
        # ===== TOTAL TIME BY ISSUE TYPE CHART =====
        time_start_row = type_end_row + 3
        ws_charts.cell(row=time_start_row - 1, column=1, value='Total Time by Issue Type Analysis')
        ws_charts.cell(row=time_start_row, column=1, value='Issue Type')
        ws_charts.cell(row=time_start_row, column=2, value='Hours')
        
        # Calculate total time by issue type
        time_by_type = Counter()
        for worklog in worklogs:
            issue_type = worklog.get('issueType', 'Unknown')
            hours = worklog.get('timeSpentHours', 0)
            time_by_type[issue_type] += hours
        
        time_data_start = time_start_row + 1
        for i, (issue_type, hours) in enumerate(sorted(time_by_type.items())):
            row = time_data_start + i
            ws_charts.cell(row=row, column=1, value=issue_type)
            ws_charts.cell(row=row, column=2, value=round(hours, 2))
        
        time_end_row = time_data_start + len(time_by_type) - 1
        
        # Create pie chart for Total Time by Issue Type
        pie_time = PieChart()
        pie_time.title = "Total Time by Issue Type"
        pie_time.width = 12
        pie_time.height = 8
        
        labels_time = Reference(ws_charts, min_col=1, min_row=time_data_start, max_row=time_end_row)
        data_time = Reference(ws_charts, min_col=2, min_row=time_data_start, max_row=time_end_row)
        
        pie_time.add_data(data_time, titles_from_data=False)
        pie_time.set_categories(labels_time)
        
        # Apply colors based on issue type configuration
        apply_colors_to_pie_chart(pie_time, sorted(time_by_type.keys()), get_issue_type_color)
        
        # Configure chart appearance - show only value and percentage
        pie_time.dataLabels = DataLabelList()
        pie_time.dataLabels.showCatName = False  # Don't show category name
        pie_time.dataLabels.showVal = True       # Show value
        pie_time.dataLabels.showPercent = True   # Show percentage
        pie_time.dataLabels.showSerName = False  # Don't show series name
        
        ws_charts.add_chart(pie_time, "D18")
        
        # ===== TIME BY AUTHOR AND ISSUE TYPE BAR CHART =====
        author_time_start_row = time_end_row + 3
        ws_charts.cell(row=author_time_start_row - 1, column=1, value='Time by Author and Issue Type Analysis')
        
        # Calculate time by author and issue type
        author_type_time = defaultdict(lambda: defaultdict(float))
        authors = set()
        issue_types = set()
        
        for worklog in worklogs:
            author = worklog.get('author', 'Unknown')
            issue_type = worklog.get('issueType', 'Unknown')
            hours = worklog.get('timeSpentHours', 0)
            
            author_type_time[author][issue_type] += hours
            authors.add(author)
            issue_types.add(issue_type)
        
        # Sort for consistent ordering
        sorted_authors = sorted(authors)
        sorted_issue_types = sorted(issue_types)
        
        # Create headers
        ws_charts.cell(row=author_time_start_row, column=1, value='Author')
        for i, issue_type in enumerate(sorted_issue_types):
            ws_charts.cell(row=author_time_start_row, column=i + 2, value=issue_type)
        ws_charts.cell(row=author_time_start_row, column=len(sorted_issue_types) + 2, value='Total')
        
        # Fill data
        author_data_start = author_time_start_row + 1
        for i, author in enumerate(sorted_authors):
            row = author_data_start + i
            ws_charts.cell(row=row, column=1, value=author)
            
            total_hours = 0
            for j, issue_type in enumerate(sorted_issue_types):
                hours = round(author_type_time[author][issue_type], 2)
                ws_charts.cell(row=row, column=j + 2, value=hours)
                total_hours += hours
            
            ws_charts.cell(row=row, column=len(sorted_issue_types) + 2, value=round(total_hours, 2))
        
        author_data_end = author_data_start + len(sorted_authors) - 1
        
        # Create bar chart for Time by Author and Issue Type
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "Time by Author and Issue Type"
        bar_chart.y_axis.title = 'Hours'
        bar_chart.x_axis.title = 'Author'
        bar_chart.width = 15
        bar_chart.height = 10
        
        # Add data series for each issue type
        for i, issue_type in enumerate(sorted_issue_types):
            data_ref = Reference(ws_charts, min_col=i + 2, min_row=author_data_start, max_row=author_data_end)
            bar_chart.add_data(data_ref, titles_from_data=False)
            
            # Set series title
            series_label = SeriesLabel()
            series_label.v = issue_type
            bar_chart.series[i].tx = series_label
        
        # Add Total series
        total_data_ref = Reference(ws_charts, min_col=len(sorted_issue_types) + 2, min_row=author_data_start, max_row=author_data_end)
        bar_chart.add_data(total_data_ref, titles_from_data=False)
        
        # Set Total series title
        total_series_label = SeriesLabel()
        total_series_label.v = "Total"
        bar_chart.series[-1].tx = total_series_label
        
        # Set categories (authors)
        cats = Reference(ws_charts, min_col=1, min_row=author_data_start, max_row=author_data_end)
        bar_chart.set_categories(cats)
        
        # Enable data labels
        bar_chart.dataLabels = DataLabelList()
        bar_chart.dataLabels.showVal = True
        
        ws_charts.add_chart(bar_chart, "P18")
        
        # Update current row for sprint charts
        current_row = author_data_end + 5
        
        # ===== NEW: STACKED BAR CHART FOR TIME BY ISSUE TYPE WITH SPRINT STACKS =====
        if worklogs:
            # Calculate time by issue type and sprint
            sprint_type_time = defaultdict(lambda: defaultdict(float))
            sprints = set()
            all_issue_types = set()
            
            for worklog in worklogs:
                sprint = worklog.get('sprint', 'N/A').strip()
                issue_type = worklog.get('issueType', 'Unknown')
                hours = worklog.get('timeSpentHours', 0)
                
                if sprint and sprint != 'N/A':
                    sprint_type_time[issue_type][sprint] += hours
                    sprints.add(sprint)
                    all_issue_types.add(issue_type)
            
            if sprints and all_issue_types:
                # Sort for consistent ordering
                sorted_sprints = sorted(sprints)
                sorted_all_issue_types = sorted(all_issue_types)
                
                # Create stacked bar chart data
                stacked_start_row = current_row
                ws_charts.cell(row=stacked_start_row - 1, column=1, value='Time by Issue Type (stacked by Cycle)')
                
                # Create headers
                ws_charts.cell(row=stacked_start_row, column=1, value='Issue Type')
                for i, sprint in enumerate(sorted_sprints):
                    ws_charts.cell(row=stacked_start_row, column=i + 2, value=sprint)
                
                # Fill data
                stacked_data_start = stacked_start_row + 1
                for i, issue_type in enumerate(sorted_all_issue_types):
                    row = stacked_data_start + i
                    ws_charts.cell(row=row, column=1, value=issue_type)
                    
                    for j, sprint in enumerate(sorted_sprints):
                        hours = round(sprint_type_time[issue_type][sprint], 2)
                        ws_charts.cell(row=row, column=j + 2, value=hours)
                
                stacked_data_end = stacked_data_start + len(sorted_all_issue_types) - 1
                
                # Create stacked bar chart (horizontal)
                stacked_bar = BarChart()
                stacked_bar.type = "bar"
                stacked_bar.grouping = "stacked"
                stacked_bar.overlap = 100
                stacked_bar.title = "Time by Issue Type (stacked by Cycle)"
                stacked_bar.y_axis.title = 'Issue Type'
                stacked_bar.x_axis.title = 'Hours'
                stacked_bar.width = 15
                stacked_bar.height = 10
                
                # Add data series for each sprint
                for i, sprint in enumerate(sorted_sprints):
                    data_ref = Reference(ws_charts, min_col=i + 2, min_row=stacked_data_start, max_row=stacked_data_end)
                    stacked_bar.add_data(data_ref, titles_from_data=False)
                    
                    # Set series title
                    series_label = SeriesLabel()
                    series_label.v = sprint
                    stacked_bar.series[i].tx = series_label
                
                # Set categories (issue types)
                cats_stacked = Reference(ws_charts, min_col=1, min_row=stacked_data_start, max_row=stacked_data_end)
                stacked_bar.set_categories(cats_stacked)
                
                # Enable data labels
                stacked_bar.dataLabels = DataLabelList()
                stacked_bar.dataLabels.showVal = True
                
                ws_charts.add_chart(stacked_bar, f"D{current_row + 5}")
                
                # Update current row
                current_row = stacked_data_end + 15
    else:
        current_row = type_end_row + 5
    
    # ===== INDIVIDUAL SPRINT CHARTS =====
    if issues_by_sprint:
        for sprint_id, sprint_data in issues_by_sprint.items():
            sprint_name = sprint_data.get('name', f'Sprint {sprint_id}')
            sprint_issues = sprint_data.get('issues', [])
            
            if not sprint_issues:
                continue
            
            # ===== SPRINT STATUS CHART =====
            ws_charts.cell(row=current_row, column=1, value=f'{sprint_name} - Issues by Status')
            ws_charts.cell(row=current_row + 1, column=1, value='Status')
            ws_charts.cell(row=current_row + 1, column=2, value='Count')
            
            # Count issues by status for this sprint
            sprint_status_counts = Counter()
            for issue in sprint_issues:
                status = issue.get('fields', {}).get('status', {}).get('name', 'Unknown')
                sprint_status_counts[status] += 1
            
            sprint_status_start = current_row + 2
            for i, (status, count) in enumerate(sorted(sprint_status_counts.items())):
                row = sprint_status_start + i
                ws_charts.cell(row=row, column=1, value=status)
                ws_charts.cell(row=row, column=2, value=count)
            
            sprint_status_end = sprint_status_start + len(sprint_status_counts) - 1
            
            # Create pie chart for this sprint's status
            pie_sprint_status = PieChart()
            pie_sprint_status.title = f"{sprint_name} - Issues by Status"
            pie_sprint_status.width = 10
            pie_sprint_status.height = 7
            
            labels_sprint_status = Reference(ws_charts, min_col=1, min_row=sprint_status_start, max_row=sprint_status_end)
            data_sprint_status = Reference(ws_charts, min_col=2, min_row=sprint_status_start, max_row=sprint_status_end)
            
            pie_sprint_status.add_data(data_sprint_status, titles_from_data=False)
            pie_sprint_status.set_categories(labels_sprint_status)
            
            # Apply colors based on status configuration
            apply_colors_to_pie_chart(pie_sprint_status, sorted(sprint_status_counts.keys()), get_status_color)
            
            # Configure chart appearance - show only value and percentage
            pie_sprint_status.dataLabels = DataLabelList()
            pie_sprint_status.dataLabels.showCatName = False  # Don't show category name
            pie_sprint_status.dataLabels.showVal = True       # Show value
            pie_sprint_status.dataLabels.showPercent = True   # Show percentage
            pie_sprint_status.dataLabels.showSerName = False  # Don't show series name
            
            # Position status chart
            chart_position_status = f"P{current_row}"
            ws_charts.add_chart(pie_sprint_status, chart_position_status)
            
            # ===== SPRINT TYPE CHART =====
            # Move to next section for type chart
            current_row = sprint_status_end + 3
            
            # Create data for this sprint's type chart
            ws_charts.cell(row=current_row, column=1, value=f'{sprint_name} - Issues by Type')
            ws_charts.cell(row=current_row + 1, column=1, value='Issue Type')
            ws_charts.cell(row=current_row + 1, column=2, value='Count')
            
            # Count issues by type for this sprint
            sprint_type_counts = Counter()
            for issue in sprint_issues:
                issue_type = issue.get('fields', {}).get('issuetype', {}).get('name', 'Unknown')
                sprint_type_counts[issue_type] += 1
            
            sprint_type_start = current_row + 2
            for i, (issue_type, count) in enumerate(sorted(sprint_type_counts.items())):
                row = sprint_type_start + i
                ws_charts.cell(row=row, column=1, value=issue_type)
                ws_charts.cell(row=row, column=2, value=count)
            
            sprint_type_end = sprint_type_start + len(sprint_type_counts) - 1
            
            # Create pie chart for this sprint's type
            pie_sprint_type = PieChart()
            pie_sprint_type.title = f"{sprint_name} - Issues by Type"
            pie_sprint_type.width = 10
            pie_sprint_type.height = 7
            
            labels_sprint_type = Reference(ws_charts, min_col=1, min_row=sprint_type_start, max_row=sprint_type_end)
            data_sprint_type = Reference(ws_charts, min_col=2, min_row=sprint_type_start, max_row=sprint_type_end)
            
            pie_sprint_type.add_data(data_sprint_type, titles_from_data=False)
            pie_sprint_type.set_categories(labels_sprint_type)
            
            # Apply colors based on issue type configuration
            apply_colors_to_pie_chart(pie_sprint_type, sorted(sprint_type_counts.keys()), get_issue_type_color)
            
            # Configure chart appearance - show only value and percentage
            pie_sprint_type.dataLabels = DataLabelList()
            pie_sprint_type.dataLabels.showCatName = False  # Don't show category name
            pie_sprint_type.dataLabels.showVal = True       # Show value
            pie_sprint_type.dataLabels.showPercent = True   # Show percentage
            pie_sprint_type.dataLabels.showSerName = False  # Don't show series name
            
            # Position type chart (next to status chart)
            chart_position_type = f"AB{current_row}"  # Column AB for type charts
            ws_charts.add_chart(pie_sprint_type, chart_position_type)
            
            # ===== NEW: SPRINT TIME BY ISSUE TYPE CHART =====
            if worklogs:
                # Move to next section for time chart
                current_row = sprint_type_end + 3
                
                # Filter worklogs for this sprint
                sprint_worklogs = [wl for wl in worklogs if wl.get('sprint', '').strip() == sprint_name.strip()]
                
                if sprint_worklogs:
                    # Create data for this sprint's time by issue type chart
                    ws_charts.cell(row=current_row, column=1, value=f'{sprint_name} - Time by Issue Type')
                    ws_charts.cell(row=current_row + 1, column=1, value='Issue Type')
                    ws_charts.cell(row=current_row + 1, column=2, value='Hours')
                    
                    # Calculate time by issue type for this sprint
                    sprint_time_by_type = Counter()
                    for worklog in sprint_worklogs:
                        issue_type = worklog.get('issueType', 'Unknown')
                        hours = worklog.get('timeSpentHours', 0)
                        sprint_time_by_type[issue_type] += hours
                    
                    sprint_time_start = current_row + 2
                    for i, (issue_type, hours) in enumerate(sorted(sprint_time_by_type.items())):
                        row = sprint_time_start + i
                        ws_charts.cell(row=row, column=1, value=issue_type)
                        ws_charts.cell(row=row, column=2, value=round(hours, 2))
                    
                    sprint_time_end = sprint_time_start + len(sprint_time_by_type) - 1
                    
                    # Create pie chart for this sprint's time by issue type
                    pie_sprint_time = PieChart()
                    pie_sprint_time.title = f"{sprint_name} - Time by Issue Type"
                    pie_sprint_time.width = 10
                    pie_sprint_time.height = 7
                    
                    labels_sprint_time = Reference(ws_charts, min_col=1, min_row=sprint_time_start, max_row=sprint_time_end)
                    data_sprint_time = Reference(ws_charts, min_col=2, min_row=sprint_time_start, max_row=sprint_time_end)
                    
                    pie_sprint_time.add_data(data_sprint_time, titles_from_data=False)
                    pie_sprint_time.set_categories(labels_sprint_time)
                    
                    # Apply colors based on issue type configuration
                    apply_colors_to_pie_chart(pie_sprint_time, sorted(sprint_time_by_type.keys()), get_issue_type_color)
                    
                    # Configure chart appearance - show only value and percentage
                    pie_sprint_time.dataLabels = DataLabelList()
                    pie_sprint_time.dataLabels.showCatName = False  # Don't show category name
                    pie_sprint_time.dataLabels.showVal = True       # Show value
                    pie_sprint_time.dataLabels.showPercent = True   # Show percentage
                    pie_sprint_time.dataLabels.showSerName = False  # Don't show series name
                    
                    # Position time chart (in a new column)
                    chart_position_time = f"AT{current_row}"  # Column AT for time charts
                    ws_charts.add_chart(pie_sprint_time, chart_position_time)
                    
                    # Update current row for next sprint
                    current_row = sprint_time_end + 5
                else:
                    # No worklogs for this sprint, just move to next position
                    current_row = sprint_type_end + 5
            else:
                # No worklogs data, just move to next position
                current_row = sprint_type_end + 5
