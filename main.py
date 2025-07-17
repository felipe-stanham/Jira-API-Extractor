import warnings
from urllib3.exceptions import NotOpenSSLWarning
# Suppress the warning at the very top of the script to ensure it's active
# before any other modules (like requests) are imported.
warnings.filterwarnings("ignore", category=NotOpenSSLWarning)

import os
import requests
import argparse
from dotenv import load_dotenv
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList
from collections import Counter

# Load environment variables from .env file
load_dotenv()

# Jira API credentials from environment variables
JIRA_API_URL = os.getenv("JIRA_API_URL")
JIRA_API_TOKEN = os.getenv("JIRA_API_TOKEN")
JIRA_USER_EMAIL = os.getenv("JIRA_USER_EMAIL")

def get_auth():
    """Returns the authentication object for Jira API requests."""
    return (JIRA_USER_EMAIL, JIRA_API_TOKEN)

def parse_adf_to_text(adf):
    """Parses an Atlassian Document Format object into a plain text string."""
    if not isinstance(adf, dict) or 'content' not in adf:
        # If it's not a valid ADF object, return it as is (might be a simple string).
        return str(adf)

    text_content = []
    for block in adf.get('content', []):
        if block.get('type') == 'paragraph':
            for item in block.get('content', []):
                if item.get('type') == 'text':
                    text_content.append(item.get('text', ''))
    return " ".join(text_content)

def get_issues_in_sprint(project_key, sprint_id):
    """Fetches all issues in a given sprint using the Agile API."""
    agile_url = f"{JIRA_API_URL}/rest/agile/1.0/sprint/{sprint_id}/issue"
    headers = {"Accept": "application/json"}
    params = {'jql': f'project = "{project_key}"', 'fields': 'summary,status,parent,issuetype', 'maxResults': 100}
    try:
        response = requests.get(agile_url, headers=headers, params=params, auth=get_auth())
        response.raise_for_status()
        return response.json().get('issues', [])
    except requests.exceptions.RequestException as e:
        return {'error': str(e)}

def get_work_logs(project_key, start_date, end_date):
    """Fetches all work log entries within a date range for a project."""
    search_url = f"{JIRA_API_URL}/rest/api/3/search"
    headers = {"Accept": "application/json"}
    jql = f'project = "{project_key}" AND worklogDate >= "{start_date}" AND worklogDate <= "{end_date}"'
    params = {'jql': jql, 'fields': 'worklog,summary,issuetype,status,sprint,closedSprints', 'maxResults': 1000}
    try:
        response = requests.get(search_url, headers=headers, params=params, auth=get_auth())
        response.raise_for_status()
        issues_with_worklogs = response.json().get('issues', [])
        all_worklogs = []
        for issue in issues_with_worklogs:
            issue_worklogs = issue.get('fields', {}).get('worklog', {}).get('worklogs', [])
            for worklog in issue_worklogs:
                worklog_started_str = worklog.get('started')
                if worklog_started_str[-3] != ':':
                    worklog_started_str = worklog_started_str[:-2] + ':' + worklog_started_str[-2:]
                worklog_date = datetime.fromisoformat(worklog_started_str).astimezone(timezone.utc)
                start_date_aware = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                end_date_aware = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                if start_date_aware <= worklog_date <= end_date_aware:
                    comment_obj = worklog.get('comment', {})
                    comment_text = parse_adf_to_text(comment_obj)
                    sprints = []
                    sprint_field = issue.get('fields', {}).get('sprint')
                    if sprint_field:
                        sprints.extend([s['name'] for s in sprint_field])
                    closed_sprints_field = issue.get('fields', {}).get('closedSprints')
                    if closed_sprints_field:
                        sprints.extend([s['name'] for s in closed_sprints_field])
                    sprint_names = '; '.join(sorted(list(set(sprints)))) if sprints else 'N/A'

                    all_worklogs.append({
                        'issueKey': issue.get('key'),
                        'summary': issue.get('fields', {}).get('summary'),
                        'issueType': issue.get('fields', {}).get('issuetype', {}).get('name', 'N/A'),
                        'status': issue.get('fields', {}).get('status', {}).get('name', 'N/A'),
                        'sprint': sprint_names,
                        'author': worklog.get('author', {}).get('displayName'),
                        'timeSpent': worklog.get('timeSpent'),
                        'timeSpentHours': round(worklog.get('timeSpentSeconds', 0) / 3600, 2),
                        'startedDate': worklog_date.strftime('%Y-%m-%d'),
                        'comment': comment_text
                    })
        return all_worklogs
    except requests.exceptions.RequestException as e:
        return {'error': str(e)}

def get_comments_in_date_range(project_key, start_date, end_date):
    """Fetches all comments created within a date range for a project."""
    search_url = f"{JIRA_API_URL}/rest/api/3/search"
    headers = {"Accept": "application/json"}
    jql = f'project = "{project_key}" AND updatedDate >= "{start_date}" AND updatedDate <= "{end_date}"'
    params = {'jql': jql, 'fields': 'comment,summary,status,parent,issuetype', 'maxResults': 1000}
    try:
        response = requests.get(search_url, headers=headers, params=params, auth=get_auth())
        response.raise_for_status()
        issues = response.json().get('issues', [])
        all_comments = []
        for issue in issues:
            comments_in_issue = issue.get('fields', {}).get('comment', {}).get('comments', [])
            for comment in comments_in_issue:
                comment_created_str = comment.get('created')
                if comment_created_str[-3] != ':':
                    comment_created_str = comment_created_str[:-2] + ':' + comment_created_str[-2:]
                comment_date = datetime.fromisoformat(comment_created_str).astimezone(timezone.utc)
                start_date_aware = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                end_date_aware = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                if start_date_aware <= comment_date <= end_date_aware:
                    comment_body = parse_adf_to_text(comment.get('body', ''))
                    all_comments.append({
                        'issueKey': issue.get('key'),
                        'summary': issue.get('fields', {}).get('summary'),
                        'status': issue.get('fields', {}).get('status', {}).get('name', 'N/A'),
                        'parent_summary': issue.get('fields', {}).get('parent', {}).get('fields', {}).get('summary', 'N/A'),
                        'issueType': issue.get('fields', {}).get('issuetype', {}).get('name', 'N/A'),
                        'comment_date': comment_date.strftime('%Y-%m-%d'),
                        'comment_author': comment.get('author', {}).get('displayName'),
                        'comment_body': comment_body
                    })
        return all_comments
    except requests.exceptions.RequestException as e:
        return {'error': str(e)}

def save_to_excel(issues, worklogs, comments):
    """Saves the fetched data to an Excel file, including a pie chart for issue status."""
    wb = Workbook()
    wb.remove(wb.active) # Remove the default sheet

    if issues:
        ws_issues = wb.create_sheet(title="Sprint Issues")
        ws_issues.append(['Key', 'Type', 'Summary', 'Status', 'Parent Summary'])
        for issue in issues:
            fields = issue.get('fields', {})
            parent_summary = fields.get('parent', {}).get('fields', {}).get('summary', 'N/A')
            issue_type = fields.get('issuetype', {}).get('name', 'N/A')
            ws_issues.append([issue.get('key'), issue_type, fields.get('summary'), fields.get('status', {}).get('name'), parent_summary])

        # --- Create Summary Table and Pie Chart for Issue Status ---
        if issues:
            # Use a set to get unique statuses to build the summary table
            unique_statuses = sorted(list(set(issue.get('fields', {}).get('status', {}).get('name', 'N/A') for issue in issues)))
            
            # Define the range of the issue data
            max_data_row = ws_issues.max_row
            status_data_range = f'D2:D{max_data_row}'

            # Create headers for the summary table
            summary_start_col = 7 # Column G
            ws_issues.cell(row=1, column=summary_start_col, value='Status')
            ws_issues.cell(row=1, column=summary_start_col + 1, value='Count')
            ws_issues.cell(row=1, column=summary_start_col + 2, value='Percentage')

            # Populate the summary table with statuses and formulas
            for i, status in enumerate(unique_statuses, start=2):
                # Status Name
                status_cell = ws_issues.cell(row=i, column=summary_start_col, value=status)
                # Count Formula
                count_formula = f'=COUNTIF({status_data_range}, "{status}")'
                count_cell = ws_issues.cell(row=i, column=summary_start_col + 1, value=count_formula)

            # Add a 'Total' row
            total_row = len(unique_statuses) + 2
            ws_issues.cell(row=total_row, column=summary_start_col, value="Total")
            total_count_cell_ref = f'H{total_row}'
            ws_issues[total_count_cell_ref] = f'=SUM(H2:H{total_row - 1})'

            # Add percentage formulas now that the total is available
            for i in range(2, total_row):
                percentage_cell = ws_issues.cell(row=i, column=summary_start_col + 2)
                percentage_cell.value = f'=H{i}/${total_count_cell_ref}'
                percentage_cell.number_format = '0.00%'

            # Create the Pie Chart
            pie = PieChart()
            pie.title = "Issues by Status"
            
            # Configure data labels to show only value and percentage
            pie.dataLabels = DataLabelList()
            pie.dataLabels.showVal = True
            pie.dataLabels.showPercent = True
            pie.dataLabels.showCatName = False
            pie.dataLabels.showSerName = False

            labels = Reference(ws_issues, min_col=summary_start_col, min_row=2, max_row=total_row - 1)
            data = Reference(ws_issues, min_col=summary_start_col + 1, min_row=2, max_row=total_row - 1)
            
            pie.add_data(data, titles_from_data=False)
            pie.set_categories(labels)

            # Add the chart to the sheet, placing it to the right of the summary table
            ws_issues.add_chart(pie, "K2")

    if worklogs:
        ws_worklogs = wb.create_sheet(title="Work Logs")
        ws_worklogs.append(['Issue Key', 'Issue Type', 'Summary', 'Status', 'Sprint', 'Author', 'Time Spent (Hours)', 'Date', 'Comment'])
        for log in worklogs:
            ws_worklogs.append([log['issueKey'], log['issueType'], log['summary'], log['status'], log['sprint'], log['author'], log['timeSpentHours'], log['startedDate'], log['comment']])

        # --- Create Summary Tables and Charts for Work Logs ---
        max_log_row = ws_worklogs.max_row
        author_range = f'F2:F{max_log_row}'
        issuetype_range = f'B2:B{max_log_row}'
        sprint_range = f'E2:E{max_log_row}'
        hours_range = f'G2:G{max_log_row}'

        # --- Table A: Hours by Author and Issue Type ---
        table_a_start_col = 11 # Column K
        unique_authors = sorted(list(set(log['author'] for log in worklogs)))
        unique_issue_types = sorted(list(set(log['issueType'] for log in worklogs)))

        ws_worklogs.cell(row=1, column=table_a_start_col, value="Author")
        for i, issue_type in enumerate(unique_issue_types):
            ws_worklogs.cell(row=1, column=table_a_start_col + 1 + i, value=issue_type)
        ws_worklogs.cell(row=1, column=table_a_start_col + 1 + len(unique_issue_types), value="Total")
        
        for r_idx, author in enumerate(unique_authors, start=2):
            ws_worklogs.cell(row=r_idx, column=table_a_start_col, value=author)
            for c_idx, issue_type in enumerate(unique_issue_types):
                formula = f'=SUMIFS({hours_range}, {author_range}, "{author}", {issuetype_range}, "{issue_type}")'
                ws_worklogs.cell(row=r_idx, column=table_a_start_col + 1 + c_idx, value=formula)
            
            # Add total formula for the row
            start_sum_col_letter = get_column_letter(table_a_start_col + 1)
            end_sum_col_letter = get_column_letter(table_a_start_col + len(unique_issue_types))
            total_formula = f'=SUM({start_sum_col_letter}{r_idx}:{end_sum_col_letter}{r_idx})'
            ws_worklogs.cell(row=r_idx, column=table_a_start_col + 1 + len(unique_issue_types), value=total_formula)

        # --- Chart for Table A ---
        chart_a = BarChart()
        chart_a.type = "col"
        chart_a.style = 10
        chart_a.title = "Hours by Author and Issue Type"
        chart_a.y_axis.title = 'Hours'
        chart_a.x_axis.title = 'Authors'
        chart_a.legend.position = 'b'

        data = Reference(ws_worklogs, min_col=table_a_start_col + 1, min_row=1, max_col=table_a_start_col + 1 + len(unique_issue_types), max_row=len(unique_authors) + 1)
        cats = Reference(ws_worklogs, min_col=table_a_start_col, min_row=2, max_row=len(unique_authors) + 1)
        chart_a.add_data(data, titles_from_data=True)
        chart_a.set_categories(cats)
        # Ensure the x-axis is treated as a category axis to display labels
        chart_a.x_axis.number_format = '@' # or 'General'
        chart_a.x_axis.majorTickMark = 'out'
        ws_worklogs.add_chart(chart_a, f"K{len(unique_authors) + 4}")

        # --- Table B: Hours by Issue Type ---
        table_b_start_row = len(unique_authors) + 15
        ws_worklogs.cell(row=table_b_start_row, column=table_a_start_col, value="Issue Type")
        ws_worklogs.cell(row=table_b_start_row, column=table_a_start_col + 1, value="Total Hours")
        for i, issue_type in enumerate(unique_issue_types, start=1):
            ws_worklogs.cell(row=table_b_start_row + i, column=table_a_start_col, value=issue_type)
            ws_worklogs.cell(row=table_b_start_row + i, column=table_a_start_col + 1, value=f'=SUMIF({issuetype_range}, "{issue_type}", {hours_range})')

        # --- Chart for Table B ---
        chart_b = PieChart()
        chart_b.title = "Hours by Issue Type"
        chart_b.dataLabels = DataLabelList(showVal=True, showPercent=True, showCatName=False, showSerName=False)
        labels_b = Reference(ws_worklogs, min_col=table_a_start_col, min_row=table_b_start_row + 1, max_row=table_b_start_row + len(unique_issue_types))
        data_b = Reference(ws_worklogs, min_col=table_a_start_col + 1, min_row=table_b_start_row + 1, max_row=table_b_start_row + len(unique_issue_types))
        chart_b.add_data(data_b, titles_from_data=False)
        chart_b.set_categories(labels_b)
        ws_worklogs.add_chart(chart_b, f"R{table_b_start_row}")

        # --- Table C: Hours by Sprint ---
        all_sprints = set()
        for log in worklogs:
            all_sprints.update(s.strip() for s in log['sprint'].split(';') if s.strip() != 'N/A')
        unique_sprints = sorted(list(all_sprints))

        table_c_start_row = table_b_start_row + len(unique_issue_types) + 3
        ws_worklogs.cell(row=table_c_start_row, column=table_a_start_col, value="Sprint")
        ws_worklogs.cell(row=table_c_start_row, column=table_a_start_col + 1, value="Total Hours")
        for i, sprint in enumerate(unique_sprints, start=1):
            ws_worklogs.cell(row=table_c_start_row + i, column=table_a_start_col, value=sprint)
            ws_worklogs.cell(row=table_c_start_row + i, column=table_a_start_col + 1, value=f'=SUMIF({sprint_range}, "*{sprint}*", {hours_range})')
        
        # --- Chart for Table C ---
        chart_c = PieChart()
        chart_c.title = "Hours by Sprint"
        chart_c.dataLabels = DataLabelList(showVal=True, showPercent=True, showCatName=False, showSerName=False)
        labels_c = Reference(ws_worklogs, min_col=table_a_start_col, min_row=table_c_start_row + 1, max_row=table_c_start_row + len(unique_sprints))
        data_c = Reference(ws_worklogs, min_col=table_a_start_col + 1, min_row=table_c_start_row + 1, max_row=table_c_start_row + len(unique_sprints))
        chart_c.add_data(data_c, titles_from_data=False)
        chart_c.set_categories(labels_c)
        ws_worklogs.add_chart(chart_c, f"R{table_c_start_row}")

    if comments:
        ws_comments = wb.create_sheet(title="Comments")
        ws_comments.append(['Issue Key', 'Summary', 'Status', 'Parent Summary', 'Issue Type', 'Comment Date', 'Comment Author', 'Comment'])
        for comment in comments:
            ws_comments.append([comment['issueKey'], comment['summary'], comment['status'], comment['parent_summary'], comment['issueType'], comment['comment_date'], comment['comment_author'], comment['comment_body']])

    if not wb.sheetnames:
        return False, None, "No data was fetched to save."

    try:
        filename = "JiraExport.xlsx"
        wb.save(filename)
        return True, filename, None
    except Exception as e:
        return False, None, str(e)

def main():
    """Main function to parse arguments and run the export process."""
    parser = argparse.ArgumentParser(description='Jira Sprint and Work Log Exporter.')
    parser.add_argument('--project', required=True, help='Jira Project Key (e.g., NG).')
    parser.add_argument('--sprint', help='Jira Sprint ID (optional).')
    parser.add_argument('--start_date', help='Start date for work logs and comments (YYYY-MM-DD, optional).')
    parser.add_argument('--end_date', help='End date for work logs and comments (YYYY-MM-DD, optional).')

    args = parser.parse_args()

    # Validate that if one date is provided, both are.
    if (args.start_date and not args.end_date) or (not args.start_date and args.end_date):
        print("Error: Both --start_date and --end_date must be provided together.")
        return

    issues = None
    if args.sprint:
        print(f"Fetching issues for sprint {args.sprint} in project {args.project}...")
        issues = get_issues_in_sprint(args.project.upper(), args.sprint)
        if isinstance(issues, dict) and 'error' in issues:
            print(f"Error fetching issues: {issues['error']}")
            return

    worklogs = None
    comments = None
    if args.start_date and args.end_date:
        print(f"Fetching work logs from {args.start_date} to {args.end_date}...")
        worklogs = get_work_logs(args.project.upper(), args.start_date, args.end_date)
        if isinstance(worklogs, dict) and 'error' in worklogs:
            print(f"Error fetching work logs: {worklogs['error']}")
            return

        print(f"Fetching comments from {args.start_date} to {args.end_date}...")
        comments = get_comments_in_date_range(args.project.upper(), args.start_date, args.end_date)
        if isinstance(comments, dict) and 'error' in comments:
            print(f"Error fetching comments: {comments['error']}")
            return

    print("Saving data to Excel...")
    success, filename, error = save_to_excel(issues, worklogs, comments)
    if success:
        summary = []
        if issues is not None:
            summary.append(f"{len(issues)} issues")
        if worklogs is not None:
            summary.append(f"{len(worklogs)} work logs")
        if comments is not None:
            summary.append(f"{len(comments)} comments")
        print(f"\nExport complete! Found {', '.join(summary)}.")
        print(f"Data saved to {os.path.abspath(filename)}")
    else:
        print(f"Error saving to Excel: {error}")

if __name__ == '__main__':
    main()
